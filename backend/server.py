from fastapi import FastAPI, APIRouter, HTTPException, Depends, Header, Request
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import json
import base64
import secrets
import logging
import bcrypt
import jwt as pyjwt
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta, date

import httpx
import stripe
from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
_raw_db = client[os.environ['DB_NAME']]
rdb = _raw_db  # handle sem scoping (sistema/público)

EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY', '')
JWT_SECRET = os.environ.get('JWT_SECRET', 'dev-secret')
JWT_ALG = "HS256"

stripe.api_key = os.environ.get('STRIPE_SECRET_KEY') or 'sk_test_emergent'
STRIPE_WEBHOOK_SECRET = os.environ.get('STRIPE_WEBHOOK_SECRET', '')

EMAIL_BASE_URL = "https://integrations.emergentagent.com"
EMERGENT_EMAIL_KEY = os.environ.get('EMERGENT_EMAIL_KEY', '')
EMAIL_FROM_NAME = os.environ.get('EMAIL_FROM_NAME', 'StudioHub AI')
PHOTOGRAPHER_EMAIL = os.environ.get('PHOTOGRAPHER_EMAIL', '')

# ---------------- Fase 5.2: Tenant Isolation (helper central único) ----------------
from contextvars import ContextVar
current_org: ContextVar = ContextVar("current_org", default=None)

# Coleções isoladas por organização (organization_id auto-injetado)
SCOPED_COLLECTIONS = {
    "clients", "sessions", "galleries", "store_orders", "products", "store_categories",
    "quotes", "contracts", "invoices", "receivables", "payables", "events",
    "notifications", "activities", "ai_messages", "settings",
}


class _TColl:
    """Wrapper de coleção: injeta organization_id em TODAS as leituras/escritas quando há org no contexto."""
    def __init__(self, coll, name):
        self._c = coll
        self._n = name

    def _on(self):
        return self._n in SCOPED_COLLECTIONS and current_org.get() is not None

    def _q(self, q):
        if self._on():
            q = dict(q or {})
            q["organization_id"] = current_org.get()
            return q
        return q if q is not None else {}

    def find(self, q=None, *a, **k):
        return self._c.find(self._q(q), *a, **k)

    async def find_one(self, q=None, *a, **k):
        return await self._c.find_one(self._q(q), *a, **k)

    async def count_documents(self, q=None, *a, **k):
        return await self._c.count_documents(self._q(q), *a, **k)

    async def distinct(self, key, q=None, *a, **k):
        return await self._c.distinct(key, self._q(q), *a, **k)

    def aggregate(self, pipeline, *a, **k):
        if self._on():
            pipeline = [{"$match": {"organization_id": current_org.get()}}] + list(pipeline)
        return self._c.aggregate(pipeline, *a, **k)

    async def insert_one(self, doc, *a, **k):
        if self._on() and isinstance(doc, dict) and "organization_id" not in doc:
            doc["organization_id"] = current_org.get()
        return await self._c.insert_one(doc, *a, **k)

    async def insert_many(self, docs, *a, **k):
        if self._on():
            for d in docs:
                if isinstance(d, dict) and "organization_id" not in d:
                    d["organization_id"] = current_org.get()
        return await self._c.insert_many(docs, *a, **k)

    async def update_one(self, q, u, *a, **k):
        return await self._c.update_one(self._q(q), u, *a, **k)

    async def update_many(self, q, u, *a, **k):
        return await self._c.update_many(self._q(q), u, *a, **k)

    async def delete_one(self, q, *a, **k):
        return await self._c.delete_one(self._q(q), *a, **k)

    async def delete_many(self, q, *a, **k):
        return await self._c.delete_many(self._q(q), *a, **k)

    async def find_one_and_update(self, q, u, *a, **k):
        return await self._c.find_one_and_update(self._q(q), u, *a, **k)

    async def create_index(self, *a, **k):
        return await self._c.create_index(*a, **k)


class _TenantDB:
    def __init__(self, raw):
        self._raw = raw

    def __getattr__(self, name):
        return _TColl(self._raw[name], name)


db = _TenantDB(_raw_db)  # handle por defeito (fail-closed: gestão exige org)

# Rotas públicas (sem token de estúdio) — não exigem organização
PUBLIC_PREFIXES = ("/api/public/", "/api/portal/", "/api/auth/")
PUBLIC_EXACT = {"/api/", "/api/seed", "/api/stripe/webhook"}
PUBLIC_GET = {"/api/settings", "/api/templates", "/api/store/products",
              "/api/store/categories", "/api/store/orders/states"}


async def tenant_context(request: Request):
    org = None
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        try:
            payload = pyjwt.decode(auth[7:], JWT_SECRET, algorithms=[JWT_ALG])
            if payload.get("role") == "studio":
                org = payload.get("organization_id")
                if not org:
                    u = await _raw_db.users.find_one({"id": payload.get("sub")})
                    org = u.get("organization_id") if u else None
        except Exception:
            org = None
    current_org.set(org)
    path = request.url.path
    is_public = (path in PUBLIC_EXACT or any(path.startswith(p) for p in PUBLIC_PREFIXES)
                 or (request.method == "GET" and path in PUBLIC_GET))
    if not is_public and org is None:
        raise HTTPException(401, "Não autenticado")


def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def create_client_token(client_id: str, email: str) -> str:
    payload = {"sub": client_id, "email": email, "role": "client",
               "exp": datetime.now(timezone.utc) + timedelta(days=7)}
    return pyjwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)

app = FastAPI(title="StudioHub AI")
api_router = APIRouter(prefix="/api", dependencies=[Depends(tenant_context)])


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def today():
    return datetime.now(timezone.utc).date()


# ---------------- Models ----------------
class Client(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: Optional[str] = ""
    phone: Optional[str] = ""
    whatsapp: Optional[str] = ""
    address: Optional[str] = ""
    tax_id: Optional[str] = ""
    postal_code: Optional[str] = ""
    region: Optional[str] = ""
    city: Optional[str] = ""
    district: Optional[str] = ""
    nif: Optional[str] = ""
    birthdate: Optional[str] = ""
    photo: Optional[str] = ""
    client_type: str = "particular"
    status: str = "ativo"
    origin: Optional[str] = "outro"
    company: Optional[str] = ""
    tags: List[str] = []
    notes: Optional[str] = ""
    favorite: bool = False
    password_hash: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class ClientCreate(BaseModel):
    name: str
    email: Optional[str] = ""
    phone: Optional[str] = ""
    whatsapp: Optional[str] = ""
    address: Optional[str] = ""
    tax_id: Optional[str] = ""
    postal_code: Optional[str] = ""
    region: Optional[str] = ""
    city: Optional[str] = ""
    district: Optional[str] = ""
    nif: Optional[str] = ""
    birthdate: Optional[str] = ""
    photo: Optional[str] = ""
    client_type: str = "particular"
    status: str = "ativo"
    origin: Optional[str] = "outro"
    company: Optional[str] = ""
    tags: List[str] = []
    notes: Optional[str] = ""
    favorite: bool = False


class Session(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    type: str = "retrato"
    client_id: Optional[str] = ""
    client_name: Optional[str] = ""
    date: Optional[str] = ""
    time: Optional[str] = ""
    location: Optional[str] = ""
    status: str = "agendada"
    photographer: Optional[str] = ""
    value: float = 0
    notes: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class SessionCreate(BaseModel):
    title: str
    type: str = "retrato"
    client_id: Optional[str] = ""
    client_name: Optional[str] = ""
    date: Optional[str] = ""
    time: Optional[str] = ""
    location: Optional[str] = ""
    status: str = "agendada"
    photographer: Optional[str] = ""
    value: float = 0
    notes: Optional[str] = ""


class Photo(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    url: str
    name: Optional[str] = ""
    category: Optional[str] = ""
    ai_score: Optional[float] = None
    ai_tags: List[str] = []
    ai_selected: bool = False
    ai_reason: Optional[str] = ""
    featured: bool = False
    stars: int = 0
    favorite: bool = False
    selected: bool = False
    client_favorite: bool = False
    client_selected: bool = False
    approval: str = "pendente"
    comments: List[dict] = []


class Gallery(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    client_name: Optional[str] = ""
    project_id: Optional[str] = ""
    session_id: Optional[str] = ""
    type: str = "sessao"
    date: Optional[str] = ""
    description: Optional[str] = ""
    cover: Optional[str] = ""
    photos: List[Photo] = []
    status: str = "pendente"
    password: Optional[str] = ""
    access_token: Optional[str] = ""
    link_expires: Optional[str] = ""
    watermark: bool = False
    categories: List[str] = []
    created_at: str = Field(default_factory=now_iso)


class GalleryCreate(BaseModel):
    title: str
    client_name: Optional[str] = ""
    project_id: Optional[str] = ""
    session_id: Optional[str] = ""
    type: str = "sessao"
    date: Optional[str] = ""
    description: Optional[str] = ""
    cover: Optional[str] = ""
    password: Optional[str] = ""
    status: str = "pendente"


class PhotoAdd(BaseModel):
    url: str
    name: Optional[str] = ""


class EventItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    client_name: Optional[str] = ""
    date: str
    time: Optional[str] = ""
    type: str = "sessao"
    location: Optional[str] = ""
    notes: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class EventCreate(BaseModel):
    title: str
    client_name: Optional[str] = ""
    date: str
    time: Optional[str] = ""
    type: str = "sessao"
    location: Optional[str] = ""
    notes: Optional[str] = ""


class InvoiceItem(BaseModel):
    description: str
    quantity: float = 1
    price: float = 0


class Invoice(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    number: str
    client_name: str
    type: str = "fatura"
    service: Optional[str] = "sessao"
    status: str = "pendente"
    issue_date: str = Field(default_factory=lambda: now_iso()[:10])
    due_date: Optional[str] = ""
    items: List[InvoiceItem] = []
    tax_rate: float = 23
    notes: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class InvoiceCreate(BaseModel):
    client_name: str
    type: str = "fatura"
    service: Optional[str] = "sessao"
    status: str = "pendente"
    due_date: Optional[str] = ""
    items: List[InvoiceItem] = []
    tax_rate: float = 23
    notes: Optional[str] = ""


class Receivable(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    project: Optional[str] = ""
    total: float = 0
    received: float = 0
    due_date: Optional[str] = ""
    method: Optional[str] = ""
    payments: List[dict] = []
    created_at: str = Field(default_factory=now_iso)


class ReceivableCreate(BaseModel):
    client_name: str
    project: Optional[str] = ""
    total: float = 0
    received: float = 0
    due_date: Optional[str] = ""
    method: Optional[str] = ""


PAYABLE_CATEGORIES = [
    "Equipamentos", "Marketing", "Publicidade", "Transporte", "Combustível",
    "Alimentação", "Freelancers", "Fotógrafos", "Designers", "Impressões",
    "Álbuns", "Fornecedores", "Software", "Assinaturas", "Impostos", "Outros",
]


class Payable(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    supplier: str = ""
    description: str = ""
    category: str = "Outros"
    amount: float = 0
    due_date: Optional[str] = ""
    paid_date: Optional[str] = ""
    method: Optional[str] = ""
    status: str = "pendente"
    notes: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class PayableCreate(BaseModel):
    supplier: str = ""
    description: str = ""
    category: str = "Outros"
    amount: float = 0
    due_date: Optional[str] = ""
    paid_date: Optional[str] = ""
    method: Optional[str] = ""
    status: str = "pendente"
    notes: Optional[str] = ""


# ---------------- Loja Online (Store) ----------------
class Product(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = ""
    category: Optional[str] = ""
    price: float = 0
    image_url: Optional[str] = ""
    sku: Optional[str] = ""
    active: bool = True
    created_at: str = Field(default_factory=now_iso)


class ProductCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    category: Optional[str] = ""
    price: float = 0
    image_url: Optional[str] = ""
    sku: Optional[str] = ""
    active: bool = True


class StoreCategory(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = ""
    active: bool = True
    created_at: str = Field(default_factory=now_iso)


ORDER_STATES = ["novo", "pago", "em_producao", "enviado", "entregue", "cancelado"]


class OrderItem(BaseModel):
    product_id: Optional[str] = ""
    name: str
    price: float = 0
    quantity: int = 1
    photo_name: Optional[str] = ""
    photo_url: Optional[str] = ""
    notes: Optional[str] = ""


class Order(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    number: str
    customer_name: Optional[str] = ""
    customer_email: Optional[str] = ""
    customer_phone: Optional[str] = ""
    items: List[OrderItem] = []
    total: float = 0
    status: str = "novo"
    notes: Optional[str] = ""
    payment_status: str = "pending"
    stripe_session_id: Optional[str] = ""
    stripe_payment_intent_id: Optional[str] = ""
    paid_at: Optional[str] = ""
    currency: Optional[str] = "eur"
    amount: Optional[float] = 0
    history: List[dict] = []
    created_at: str = Field(default_factory=now_iso)


class OrderCreate(BaseModel):
    customer_name: Optional[str] = ""
    customer_email: Optional[str] = ""
    customer_phone: Optional[str] = ""
    items: List[OrderItem] = []
    notes: Optional[str] = ""
    status: Optional[str] = "novo"


class Quote(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    number: str
    client_name: str
    title: str = "Proposta fotográfica"
    status: str = "rascunho"
    items: List[InvoiceItem] = []
    tax_rate: float = 23
    valid_until: Optional[str] = ""
    notes: Optional[str] = ""
    template: Optional[str] = "personalizado"
    contract_id: Optional[str] = ""
    invoice_id: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class QuoteCreate(BaseModel):
    client_name: str
    title: str = "Proposta fotográfica"
    status: str = "rascunho"
    items: List[InvoiceItem] = []
    tax_rate: float = 23
    valid_until: Optional[str] = ""
    notes: Optional[str] = ""
    template: Optional[str] = "personalizado"


class Contract(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    number: str
    client_name: str
    title: str = "Contrato de prestação de serviços"
    body: str = ""
    status: str = "rascunho"
    signer_name: Optional[str] = ""
    signed_at: Optional[str] = ""
    template: Optional[str] = "personalizado"
    quote_id: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class ContractCreate(BaseModel):
    client_name: str
    title: str = "Contrato de prestação de serviços"
    body: str = ""
    template: Optional[str] = "personalizado"
    quote_id: Optional[str] = ""


class AiChatIn(BaseModel):
    message: str
    session_id: Optional[str] = ""


class Settings(BaseModel):
    company_name: str = "StudioHub AI"
    country: str = "PT"
    language: str = "pt"
    currency: str = "EUR"
    locale: str = "pt-PT"
    timezone: str = "Europe/Lisbon"
    date_format: str = "dd/MM/yyyy"
    tax_rate: float = 23
    tax_name: str = "NIF"
    tax_label: str = "IVA"
    address_labels: dict = Field(default_factory=lambda: {"postal_code": "Código Postal", "region": "Distrito", "city": "Concelho", "district": "Freguesia"})


DEFAULT_SETTINGS = Settings().model_dump()


# ---------------- Helpers ----------------
def clean(doc):
    doc.pop("_id", None)
    return doc


def invoice_totals(inv: dict):
    subtotal = sum(i.get("quantity", 1) * i.get("price", 0) for i in inv.get("items", []))
    tax = subtotal * inv.get("tax_rate", 0) / 100
    inv["subtotal"] = round(subtotal, 2)
    inv["tax"] = round(tax, 2)
    inv["total"] = round(subtotal + tax, 2)
    return inv


# ---------------- Clients ----------------
@api_router.get("/clients")
async def list_clients():
    docs = await db.clients.find({}, {"_id": 0, "password_hash": 0}).sort("created_at", -1).to_list(1000)
    return docs


@api_router.get("/clients/{client_id}")
async def get_client(client_id: str):
    doc = await db.clients.find_one({"id": client_id}, {"_id": 0, "password_hash": 0})
    if not doc:
        raise HTTPException(404, "Cliente não encontrado")
    return doc


@api_router.post("/clients", response_model=Client)
async def create_client(payload: ClientCreate):
    obj = Client(**payload.model_dump())
    await db.clients.insert_one(obj.model_dump())
    return obj


@api_router.put("/clients/{client_id}", response_model=Client)
async def update_client(client_id: str, payload: ClientCreate):
    if not await db.clients.find_one({"id": client_id}):
        raise HTTPException(404, "Cliente não encontrado")
    await db.clients.update_one({"id": client_id}, {"$set": payload.model_dump()})
    return await db.clients.find_one({"id": client_id}, {"_id": 0})


@api_router.patch("/clients/{client_id}/favorite", response_model=Client)
async def toggle_favorite(client_id: str):
    doc = await db.clients.find_one({"id": client_id})
    if not doc:
        raise HTTPException(404, "Cliente não encontrado")
    await db.clients.update_one({"id": client_id}, {"$set": {"favorite": not doc.get("favorite", False)}})
    return await db.clients.find_one({"id": client_id}, {"_id": 0})


@api_router.delete("/clients/{client_id}")
async def delete_client(client_id: str):
    await db.clients.delete_one({"id": client_id})
    return {"ok": True}


# ---------------- Sessions ----------------
@api_router.get("/sessions", response_model=List[Session])
async def list_sessions():
    return await db.sessions.find({}, {"_id": 0}).sort("date", 1).to_list(2000)


@api_router.post("/sessions", response_model=Session)
async def create_session(payload: SessionCreate):
    obj = Session(**payload.model_dump())
    await db.sessions.insert_one(obj.model_dump())
    return obj


@api_router.put("/sessions/{session_id}", response_model=Session)
async def update_session(session_id: str, payload: SessionCreate):
    if not await db.sessions.find_one({"id": session_id}):
        raise HTTPException(404, "Sessão não encontrada")
    await db.sessions.update_one({"id": session_id}, {"$set": payload.model_dump()})
    return await db.sessions.find_one({"id": session_id}, {"_id": 0})


@api_router.patch("/sessions/{session_id}/status", response_model=Session)
async def update_session_status(session_id: str, body: dict):
    await db.sessions.update_one({"id": session_id}, {"$set": {"status": body.get("status", "agendada")}})
    doc = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Sessão não encontrada")
    return doc


@api_router.delete("/sessions/{session_id}")
async def delete_session(session_id: str):
    await db.sessions.delete_one({"id": session_id})
    return {"ok": True}


# ---------------- Galleries ----------------
@api_router.get("/galleries", response_model=List[Gallery])
async def list_galleries():
    return await db.galleries.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)


@api_router.get("/galleries/{gallery_id}", response_model=Gallery)
async def get_gallery(gallery_id: str):
    doc = await db.galleries.find_one({"id": gallery_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Galeria não encontrada")
    return doc


@api_router.post("/galleries", response_model=Gallery)
async def create_gallery(payload: GalleryCreate):
    obj = Gallery(**payload.model_dump())
    await db.galleries.insert_one(obj.model_dump())
    return obj


@api_router.patch("/galleries/{gallery_id}/status", response_model=Gallery)
async def update_gallery_status(gallery_id: str, body: dict):
    await db.galleries.update_one({"id": gallery_id}, {"$set": {"status": body.get("status", "pendente")}})
    doc = await db.galleries.find_one({"id": gallery_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Galeria não encontrada")
    return doc


@api_router.post("/galleries/{gallery_id}/photos", response_model=Gallery)
async def add_photo(gallery_id: str, payload: PhotoAdd):
    doc = await db.galleries.find_one({"id": gallery_id})
    if not doc:
        raise HTTPException(404, "Galeria não encontrada")
    photo = Photo(url=payload.url, name=payload.name)
    update = {"$push": {"photos": photo.model_dump()}}
    if not doc.get("cover"):
        update["$set"] = {"cover": payload.url}
    await db.galleries.update_one({"id": gallery_id}, update)
    return await db.galleries.find_one({"id": gallery_id}, {"_id": 0})


@api_router.delete("/galleries/{gallery_id}/photos/{photo_id}", response_model=Gallery)
async def delete_photo(gallery_id: str, photo_id: str):
    await db.galleries.update_one({"id": gallery_id}, {"$pull": {"photos": {"id": photo_id}}})
    return await db.galleries.find_one({"id": gallery_id}, {"_id": 0})


@api_router.delete("/galleries/{gallery_id}")
async def delete_gallery(gallery_id: str):
    await db.galleries.delete_one({"id": gallery_id})
    return {"ok": True}


@api_router.post("/galleries/{gallery_id}/ai-select", response_model=Gallery)
async def ai_select(gallery_id: str):
    doc = await db.galleries.find_one({"id": gallery_id})
    if not doc:
        raise HTTPException(404, "Galeria não encontrada")
    photos = doc.get("photos", [])
    if not photos:
        raise HTTPException(400, "Sem fotos para analisar")

    chat = LlmChat(
        api_key=EMERGENT_LLM_KEY,
        session_id=f"ai-select-{gallery_id}-{uuid.uuid4()}",
        system_message=(
            "És um assistente de curadoria fotográfica profissional. Avalias fotografias "
            "quanto a composição, nitidez, iluminação, emoção e apelo comercial. "
            "Respondes SEMPRE apenas em JSON válido em português de Portugal."
        ),
    ).with_model("openai", "gpt-5.4")

    updated_photos = []
    async with httpx.AsyncClient(timeout=20) as http:
        for p in photos:
            url = p.get("url", "")
            try:
                b64 = None
                if url.startswith("data:"):
                    b64 = url.split(",", 1)[1]
                elif url.startswith("http"):
                    r = await http.get(url)
                    if r.status_code == 200:
                        b64 = base64.b64encode(r.content).decode()
                content = [ImageContent(image_base64=b64)] if b64 else []
                msg = UserMessage(
                    text=(
                        "Avalia esta fotografia. Devolve JSON com exatamente estas chaves: "
                        '{"score": (0-100 inteiro), "tags": [3 etiquetas curtas em português], '
                        '"reason": "uma frase curta a justificar a nota"}. '
                        "Sê criterioso: nitidez, composição, luz e emoção."
                    ),
                    file_contents=content,
                )
                resp = await chat.send_message(msg)
                text = (resp if isinstance(resp, str) else str(resp)).strip()
                text = text.replace("```json", "").replace("```", "").strip()
                data = json.loads(text)
                p["ai_score"] = float(data.get("score", 0))
                p["ai_tags"] = data.get("tags", [])[:3]
                p["ai_reason"] = data.get("reason", "")
            except Exception as e:
                logging.warning(f"AI photo eval failed: {e}")
                p["ai_score"] = p.get("ai_score") or 50.0
                p["ai_tags"] = p.get("ai_tags") or ["por avaliar"]
                p["ai_reason"] = p.get("ai_reason") or "Não foi possível analisar automaticamente."
            updated_photos.append(p)

    ranked = sorted(updated_photos, key=lambda x: x.get("ai_score", 0), reverse=True)
    top_ids = {ranked[i]["id"] for i in range(max(1, round(len(ranked) * 0.4)))}
    for p in updated_photos:
        p["ai_selected"] = p["id"] in top_ids

    await db.galleries.update_one({"id": gallery_id}, {"$set": {"photos": updated_photos}})
    return await db.galleries.find_one({"id": gallery_id}, {"_id": 0})


# ---------------- Premium Gallery: settings, sharing, AI search, featured ----------------
class GallerySettings(BaseModel):
    password: Optional[str] = None
    watermark: Optional[bool] = None
    link_expires: Optional[str] = None
    categories: Optional[List[str]] = None


@api_router.patch("/galleries/{gallery_id}/settings", response_model=Gallery)
async def gallery_settings(gallery_id: str, payload: GallerySettings):
    if not await db.galleries.find_one({"id": gallery_id}):
        raise HTTPException(404, "Galeria não encontrada")
    upd = {k: v for k, v in payload.model_dump().items() if v is not None}
    if upd:
        await db.galleries.update_one({"id": gallery_id}, {"$set": upd})
    return await db.galleries.find_one({"id": gallery_id}, {"_id": 0})


@api_router.post("/galleries/{gallery_id}/share", response_model=Gallery)
async def gallery_share(gallery_id: str):
    doc = await db.galleries.find_one({"id": gallery_id})
    if not doc:
        raise HTTPException(404, "Galeria não encontrada")
    token = doc.get("access_token") or uuid.uuid4().hex[:12]
    await db.galleries.update_one({"id": gallery_id}, {"$set": {"access_token": token, "status": "partilhada"}})
    return await db.galleries.find_one({"id": gallery_id}, {"_id": 0})


@api_router.patch("/galleries/{gallery_id}/photos/{photo_id}/feature", response_model=Gallery)
async def feature_photo(gallery_id: str, photo_id: str):
    doc = await db.galleries.find_one({"id": gallery_id})
    if not doc:
        raise HTTPException(404, "Galeria não encontrada")
    photos = doc.get("photos", [])
    for p in photos:
        if p["id"] == photo_id:
            p["featured"] = not p.get("featured", False)
    await db.galleries.update_one({"id": gallery_id}, {"$set": {"photos": photos}})
    return await db.galleries.find_one({"id": gallery_id}, {"_id": 0})


@api_router.patch("/galleries/{gallery_id}/photos/{photo_id}/rate", response_model=Gallery)
async def rate_photo(gallery_id: str, photo_id: str, body: dict):
    doc = await db.galleries.find_one({"id": gallery_id})
    if not doc:
        raise HTTPException(404, "Galeria não encontrada")
    stars = max(0, min(5, int(body.get("stars", 0))))
    for p in doc.get("photos", []):
        if p["id"] == photo_id:
            p["stars"] = stars
    await db.galleries.update_one({"id": gallery_id}, {"$set": {"photos": doc["photos"]}})
    return await db.galleries.find_one({"id": gallery_id}, {"_id": 0})


@api_router.patch("/galleries/{gallery_id}/photos/{photo_id}/toggle", response_model=Gallery)
async def toggle_photo(gallery_id: str, photo_id: str, body: dict):
    field = body.get("field")
    if field not in ("favorite", "selected"):
        raise HTTPException(400, "Campo inválido")
    doc = await db.galleries.find_one({"id": gallery_id})
    if not doc:
        raise HTTPException(404, "Galeria não encontrada")
    for p in doc.get("photos", []):
        if p["id"] == photo_id:
            p[field] = not p.get(field, False)
    await db.galleries.update_one({"id": gallery_id}, {"$set": {"photos": doc["photos"]}})
    return await db.galleries.find_one({"id": gallery_id}, {"_id": 0})


@api_router.post("/galleries/{gallery_id}/photos/{photo_id}/comment", response_model=Gallery)
async def comment_photo(gallery_id: str, photo_id: str, body: dict):
    text = (body.get("text") or "").strip()
    if not text:
        raise HTTPException(400, "Comentário vazio")
    doc = await db.galleries.find_one({"id": gallery_id})
    if not doc:
        raise HTTPException(404, "Galeria não encontrada")
    for p in doc.get("photos", []):
        if p["id"] == photo_id:
            p.setdefault("comments", []).append({"author": body.get("author", "Fotógrafo"), "text": text, "ts": now_iso()})
    await db.galleries.update_one({"id": gallery_id}, {"$set": {"photos": doc["photos"]}})
    return await db.galleries.find_one({"id": gallery_id}, {"_id": 0})


async def _ensure_analyzed(gallery_id):
    """Analyze photos that lack ai_tags so AI search has data."""
    doc = await db.galleries.find_one({"id": gallery_id})
    photos = doc.get("photos", [])
    if any(not p.get("ai_tags") for p in photos):
        await ai_select(gallery_id)
        doc = await db.galleries.find_one({"id": gallery_id})
    return doc


@api_router.post("/galleries/{gallery_id}/ai-search")
async def gallery_ai_search(gallery_id: str, body: dict):
    query = (body.get("query") or "").strip()
    if not query:
        raise HTTPException(400, "Consulta vazia")
    doc = await _ensure_analyzed(gallery_id)
    if not doc:
        raise HTTPException(404, "Galeria não encontrada")
    photos = doc.get("photos", [])
    catalog = [{"id": p["id"], "name": p.get("name", ""), "tags": p.get("ai_tags", []), "desc": p.get("ai_reason", "")} for p in photos]
    system = ("És um motor de pesquisa de fotografias. Recebes uma consulta em linguagem natural e uma lista de fotografias "
              "com etiquetas e descrições. Devolves APENAS JSON: {\"ids\": [ids das fotos que correspondem]}. Sem texto extra.")
    try:
        chat = LlmChat(api_key=EMERGENT_LLM_KEY, session_id=f"search-{gallery_id}-{uuid.uuid4()}", system_message=system).with_model("openai", "gpt-5.4")
        msg = UserMessage(text=f"Consulta: {query}\n\nFotografias:\n{json.dumps(catalog, ensure_ascii=False)}")
        resp = await chat.send_message(msg)
        text = (resp if isinstance(resp, str) else str(resp)).replace("```json", "").replace("```", "").strip()
        ids = json.loads(text).get("ids", [])
    except Exception as e:
        logging.warning(f"AI search failed: {e}")
        ids = [p["id"] for p in photos if any(query.lower() in t.lower() for t in p.get("ai_tags", []))]
    return {"ids": ids, "query": query}


# ---------------- Session -> Gallery ----------------
@api_router.post("/sessions/{session_id}/gallery", response_model=Gallery)
async def session_gallery(session_id: str):
    s = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not s:
        raise HTTPException(404, "Sessão não encontrada")
    existing = await db.galleries.find_one({"session_id": session_id}, {"_id": 0})
    if existing:
        return existing
    g = Gallery(title=f"{s['title']} — Galeria", client_name=s.get("client_name", ""), session_id=session_id)
    await db.galleries.insert_one(g.model_dump())
    return g.model_dump()


# ---------------- Public Client Gallery ----------------
async def _get_by_token(token):
    doc = await db.galleries.find_one({"access_token": token}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Galeria não encontrada")
    if doc.get("link_expires"):
        try:
            if doc["link_expires"] < today().isoformat():
                raise HTTPException(410, "O link expirou")
        except HTTPException:
            raise
        except Exception:
            pass
    return doc


def _watermarked(doc):
    return doc


@api_router.get("/public/galleries/{token}")
async def public_gallery(token: str):
    doc = await _get_by_token(token)
    if doc.get("password"):
        return {"protected": True, "title": doc["title"], "client_name": doc.get("client_name", "")}
    return doc


@api_router.post("/public/galleries/{token}/verify")
async def public_gallery_verify(token: str, body: dict):
    doc = await _get_by_token(token)
    if doc.get("password") and body.get("password") != doc["password"]:
        raise HTTPException(403, "Palavra-passe incorreta")
    return doc


async def _token_gallery_or_403(token, pin):
    doc = await _get_by_token(token)
    if doc.get("password") and pin != doc["password"]:
        raise HTTPException(403, "Acesso negado")
    return doc


@api_router.patch("/public/galleries/{token}/photos/{photo_id}")
async def public_photo_action(token: str, photo_id: str, body: dict):
    doc = await _token_gallery_or_403(token, body.get("pin", ""))
    action = body.get("action")
    photos = doc.get("photos", [])
    for p in photos:
        if p["id"] == photo_id:
            if action == "favorite":
                p["client_favorite"] = not p.get("client_favorite", False)
            elif action == "select":
                p["client_selected"] = not p.get("client_selected", False)
            elif action == "approve":
                p["approval"] = "aprovada"
            elif action == "reject":
                p["approval"] = "rejeitada"
    await db.galleries.update_one({"id": doc["id"]}, {"$set": {"photos": photos}})
    return await db.galleries.find_one({"id": doc["id"]}, {"_id": 0})


@api_router.post("/public/galleries/{token}/photos/{photo_id}/comment")
async def public_photo_comment(token: str, photo_id: str, body: dict):
    doc = await _token_gallery_or_403(token, body.get("pin", ""))
    text = (body.get("text") or "").strip()
    if not text:
        raise HTTPException(400, "Comentário vazio")
    photos = doc.get("photos", [])
    for p in photos:
        if p["id"] == photo_id:
            p.setdefault("comments", []).append({"author": body.get("author", "Cliente"), "text": text, "ts": now_iso()})
    await db.galleries.update_one({"id": doc["id"]}, {"$set": {"photos": photos}})
    return await db.galleries.find_one({"id": doc["id"]}, {"_id": 0})


@api_router.post("/public/galleries/{token}/order")
async def public_order(token: str, body: dict):
    doc = await _get_by_token(token)
    order = {"id": str(uuid.uuid4()), "gallery_id": doc["id"], "client_name": doc.get("client_name", ""),
             "items": body.get("items", []), "total": body.get("total", 0), "status": "recebida", "created_at": now_iso()}
    await db.orders.insert_one(dict(order))
    order.pop("_id", None)
    return {"ok": True, "order": order, "mock": True}


# ---------------- Fase 4: Automação Pós-Venda (helpers) ----------------
def fmt_eur(v):
    try:
        return f"{float(v or 0):.2f} €"
    except Exception:
        return "0.00 €"


async def send_email_raw(to: str, subject: str, html: str) -> bool:
    if not EMERGENT_EMAIL_KEY or not to:
        return False
    try:
        async with httpx.AsyncClient(timeout=30) as c:
            r = await c.post(f"{EMAIL_BASE_URL}/api/v1/email/send",
                             headers={"X-Email-Key": EMERGENT_EMAIL_KEY},
                             json={"to": [to], "subject": subject, "html": html, "from_name": EMAIL_FROM_NAME})
        r.raise_for_status()
        return True
    except Exception as e:
        logging.error(f"[EMAIL] falha ao enviar para {to}: {e}")
        return False


async def create_notification(ntype: str, title: str, message: str, order_id: str = ""):
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()), "type": ntype, "title": title, "message": message,
        "order_id": order_id, "read": False, "created_at": now_iso(),
    })


async def create_activity(atype: str, message: str, client_name: str = "", order_id: str = ""):
    await db.activities.insert_one({
        "id": str(uuid.uuid4()), "type": atype, "message": message,
        "client_name": client_name, "order_id": order_id, "created_at": now_iso(),
    })


async def order_history(oid: str, message: str):
    await db.store_orders.update_one({"id": oid}, {"$push": {"history": {"ts": now_iso(), "message": message}}})


def _order_email_customer(order):
    rows = "".join(
        f"<tr><td style='padding:6px 0;color:#333'>{it.get('quantity',1)}× {it.get('name','')}</td>"
        f"<td style='padding:6px 0;text-align:right;color:#333'>{fmt_eur((it.get('price',0) or 0)*(it.get('quantity',1) or 1))}</td></tr>"
        for it in order.get("items", []))
    return f"""
    <div style="font-family:Arial,sans-serif;max-width:560px;margin:0 auto">
      <h2 style="color:#111">Pagamento recebido</h2>
      <p style="color:#333">Obrigado pela sua compra. O seu pedido foi confirmado.</p>
      <table style="width:100%;border-collapse:collapse;margin-top:16px">
        <tr><td style="color:#888">Número</td><td style="text-align:right;color:#111"><b>{order.get('number','')}</b></td></tr>
        <tr><td style="color:#888">Data</td><td style="text-align:right;color:#111">{(order.get('paid_at') or '')[:10]}</td></tr>
      </table>
      <hr style="border:none;border-top:1px solid #eee;margin:16px 0">
      <table style="width:100%;border-collapse:collapse">{rows}</table>
      <hr style="border:none;border-top:1px solid #eee;margin:16px 0">
      <table style="width:100%"><tr><td style="color:#111"><b>Total</b></td><td style="text-align:right;color:#111"><b>{fmt_eur(order.get('total',0))}</b></td></tr></table>
    </div>"""


def _order_email_photographer(order):
    return f"""
    <div style="font-family:Arial,sans-serif;max-width:560px;margin:0 auto">
      <h2 style="color:#111">Novo pedido pago</h2>
      <table style="width:100%;border-collapse:collapse;margin-top:12px">
        <tr><td style="color:#888">Cliente</td><td style="text-align:right;color:#111">{order.get('customer_name','—')}</td></tr>
        <tr><td style="color:#888">Valor</td><td style="text-align:right;color:#111"><b>{fmt_eur(order.get('total',0))}</b></td></tr>
        <tr><td style="color:#888">Galeria</td><td style="text-align:right;color:#111">{order.get('gallery_title','—')}</td></tr>
        <tr><td style="color:#888">Pedido</td><td style="text-align:right;color:#111">{order.get('number','')}</td></tr>
      </table>
    </div>"""


async def run_paid_automation(session_id: str):
    order = await db.store_orders.find_one({"stripe_session_id": session_id})
    if not order:
        return
    # Idempotência atómica: só corre uma vez por pedido
    res = await db.store_orders.update_one(
        {"id": order["id"], "automation_done": {"$ne": True}}, {"$set": {"automation_done": True}})
    if res.modified_count == 0:
        return
    current_org.set(order.get("organization_id"))
    oid = order["id"]; number = order.get("number", ""); total = order.get("total", 0) or 0
    cname = order.get("customer_name") or "Cliente"
    paid_at = order.get("paid_at") or now_iso()
    await order_history(oid, "Pagamento confirmado pelo Stripe.")
    # 1) Receita automática (Financeiro)
    rec = Receivable(client_name=cname, project=f"Pedido Online {number}", total=total, received=total,
                     method="Stripe", due_date=paid_at[:10])
    rdoc = rec.model_dump()
    rdoc.update({"origin": "Pedido Online", "order_number": number, "order_id": oid,
                 "paid_date": paid_at, "type": "receita"})
    await db.receivables.insert_one(rdoc)
    await order_history(oid, "Receita criada automaticamente.")
    # 2) Email ao cliente
    if order.get("customer_email"):
        ok = await send_email_raw(order["customer_email"], "Pagamento recebido", _order_email_customer(order))
        await order_history(oid, "Email enviado ao cliente." if ok else "Falha ao enviar email ao cliente.")
    else:
        await order_history(oid, "Cliente sem email. Email de confirmação não enviado.")
    # 3) Email ao fotógrafo
    ok2 = await send_email_raw(PHOTOGRAPHER_EMAIL, "Novo pedido pago", _order_email_photographer(order))
    await order_history(oid, "Email enviado ao fotógrafo." if ok2 else "Falha ao enviar email ao fotógrafo.")
    # 4) Notificação
    await create_notification("payment_received", "Pagamento recebido", f"Pedido {number} pago — {fmt_eur(total)}", oid)
    await order_history(oid, "Notificação criada.")
    # 5) CRM
    await create_activity("payment", f"Pedido {number} pago.", cname, oid)
    await order_history(oid, "CRM atualizado.")


@api_router.post("/public/galleries/{token}/store-order")
async def public_store_order(token: str, body: dict):
    doc = await _get_by_token(token)
    current_org.set(doc.get("organization_id"))
    items = body.get("items", [])
    if not items:
        raise HTTPException(400, "Carrinho vazio")
    # Recompute prices server-side from products (fonte da verdade — evita manipulação)
    prod_ids = [it.get("product_id") for it in items if it.get("product_id")]
    prods = {}
    if prod_ids:
        async for p in db.products.find({"id": {"$in": prod_ids}}):
            prods[p["id"]] = p
    safe_items = []
    for it in items:
        p = prods.get(it.get("product_id"))
        price = float(p["price"]) if p else float(it.get("price", 0) or 0)
        name = p["name"] if p else it.get("name", "Produto")
        qty = max(1, int(it.get("quantity", 1) or 1))
        safe_items.append({
            "product_id": it.get("product_id", ""), "name": name, "price": price, "quantity": qty,
            "photo_name": it.get("photo_name", ""), "photo_url": it.get("photo_url", ""), "notes": it.get("notes", ""),
        })
    total = order_total(safe_items)
    if total <= 0:
        raise HTTPException(400, "Total inválido")
    count = await db.store_orders.count_documents({})
    number = f"ENC-{datetime.now().year}-{count + 1:04d}"
    obj = Order(number=number, customer_name=body.get("customer_name", ""), customer_email=body.get("customer_email", ""),
                customer_phone=body.get("customer_phone", ""), items=safe_items, total=total,
                status="novo", notes=body.get("notes", ""), payment_status="pending", currency="eur", amount=total)
    order = obj.model_dump()
    order["gallery_token"] = token
    order["gallery_id"] = doc.get("id", "")
    order["gallery_title"] = doc.get("title", "")
    photos, seen = [], set()
    for it in safe_items:
        u = it.get("photo_url", "")
        if u and u not in seen:
            seen.add(u)
            photos.append({"name": it.get("photo_name", ""), "url": u})
    order["photos"] = photos

    origin = (body.get("origin_url") or "").rstrip("/")
    line_items = [{
        "price_data": {
            "currency": "eur",
            "unit_amount": int(round(it["price"] * 100)),
            "product_data": {"name": (it["name"] + (f" — {it['photo_name']}" if it.get("photo_name") else ""))[:250]},
        },
        "quantity": it["quantity"],
    } for it in safe_items]
    try:
        session = stripe.checkout.Session.create(
            mode="payment",
            line_items=line_items,
            success_url=f"{origin}/payment/success?session_id={{CHECKOUT_SESSION_ID}}",
            cancel_url=f"{origin}/payment/cancel?order={order['id']}",
            customer_email=(order.get("customer_email") or None),
            metadata={"order_id": order["id"], "order_number": number, "gallery_id": order["gallery_id"]},
        )
    except Exception as e:
        logger.exception("Stripe checkout error")
        raise HTTPException(502, f"Erro ao criar sessão de pagamento: {e}")
    order["stripe_session_id"] = session.id
    order["history"] = [
        {"ts": now_iso(), "message": "Pedido criado."},
        {"ts": now_iso(), "message": "Checkout iniciado."},
    ]
    await db.store_orders.insert_one(order)
    await create_notification("new_order", "Novo pedido", f"Pedido {number} criado — {fmt_eur(total)}", order["id"])
    await create_activity("order", f"Pedido {number} criado.", order.get("customer_name") or "Cliente", order["id"])
    return {"ok": True, "order": clean(order), "checkout_url": session.url, "session_id": session.id}


@api_router.get("/public/checkout/status/{session_id}")
async def public_checkout_status(session_id: str):
    # Apenas para atualização da interface. A confirmação oficial é feita pelo webhook.
    order = await db.store_orders.find_one({"stripe_session_id": session_id}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Pedido não encontrado")
    live = None
    try:
        s = stripe.checkout.Session.retrieve(session_id)
        live = s.payment_status
    except Exception:
        pass
    return {"session_id": session_id, "order_number": order.get("number"),
            "payment_status": order.get("payment_status", "pending"),
            "stripe_status": live, "status": order.get("status"),
            "total": order.get("total"), "currency": order.get("currency", "eur")}


@api_router.post("/stripe/webhook")
async def stripe_webhook(request: Request):
    payload = await request.body()
    sig = request.headers.get("stripe-signature", "")
    try:
        event = stripe.Webhook.construct_event(payload, sig, STRIPE_WEBHOOK_SECRET)
    except Exception:
        raise HTTPException(400, "Invalid signature")
    obj = event["data"]["object"]
    etype = event["type"]
    now = now_iso()
    if etype == "checkout.session.completed":
        await db.store_orders.update_one(
            {"stripe_session_id": obj["id"], "payment_status": {"$ne": "paid"}},
            {"$set": {"payment_status": "paid", "stripe_payment_intent_id": obj.get("payment_intent", ""), "paid_at": now}})
        await run_paid_automation(obj["id"])
    elif etype == "checkout.session.async_payment_succeeded":
        await db.store_orders.update_one(
            {"stripe_session_id": obj["id"], "payment_status": {"$ne": "paid"}},
            {"$set": {"payment_status": "paid", "paid_at": now}})
        await run_paid_automation(obj["id"])
    elif etype in ("checkout.session.async_payment_failed", "checkout.session.expired"):
        await db.store_orders.update_one(
            {"stripe_session_id": obj["id"]}, {"$set": {"payment_status": "failed"}})
        o = await db.store_orders.find_one({"stripe_session_id": obj["id"]})
        if o and not o.get("payment_failed_notified"):
            await db.store_orders.update_one({"id": o["id"]}, {"$set": {"payment_failed_notified": True}})
            await order_history(o["id"], "Pagamento não concluído (falhou ou expirou).")
            await create_notification("order_cancelled", "Pedido cancelado", f"Pagamento não concluído — Pedido {o.get('number','')}", o["id"])
    elif etype == "charge.refunded":
        await db.store_orders.update_one(
            {"stripe_payment_intent_id": obj.get("payment_intent")}, {"$set": {"payment_status": "refunded"}})
    return {"status": "ok"}


# ---------------- Calendar Events ----------------
@api_router.get("/events", response_model=List[EventItem])
async def list_events():
    return await db.events.find({}, {"_id": 0}).sort("date", 1).to_list(1000)


@api_router.post("/events", response_model=EventItem)
async def create_event(payload: EventCreate):
    obj = EventItem(**payload.model_dump())
    await db.events.insert_one(obj.model_dump())
    return obj


@api_router.delete("/events/{event_id}")
async def delete_event(event_id: str):
    await db.events.delete_one({"id": event_id})
    return {"ok": True}


# ---------------- Invoices ----------------
@api_router.get("/invoices")
async def list_invoices():
    docs = await db.invoices.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [invoice_totals(d) for d in docs]


@api_router.post("/invoices")
async def create_invoice(payload: InvoiceCreate):
    count = await db.invoices.count_documents({})
    number = f"{datetime.now().year}-{count + 1:04d}"
    obj = Invoice(number=number, **payload.model_dump())
    doc = obj.model_dump()
    await db.invoices.insert_one(doc)
    return invoice_totals(clean(doc))


@api_router.put("/invoices/{invoice_id}/status")
async def update_invoice_status(invoice_id: str, body: dict):
    await db.invoices.update_one({"id": invoice_id}, {"$set": {"status": body.get("status", "pendente")}})
    doc = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Fatura não encontrada")
    return invoice_totals(doc)


@api_router.delete("/invoices/{invoice_id}")
async def delete_invoice(invoice_id: str):
    await db.invoices.delete_one({"id": invoice_id})
    return {"ok": True}


# ---------------- Contas a Receber (Receivables) ----------------
def receivable_view(r: dict):
    total = round(r.get("total", 0) or 0, 2)
    received = round(r.get("received", 0) or 0, 2)
    balance = round(total - received, 2)
    if balance < 0:
        balance = 0
    due = r.get("due_date", "") or ""
    overdue = False
    if due:
        try:
            overdue = date.fromisoformat(due[:10]) < today()
        except Exception:
            overdue = False
    if total > 0 and received >= total:
        status = "pago"
    elif received > 0:
        status = "parcial"
    elif overdue:
        status = "vencido"
    else:
        status = "pendente"
    r["balance"] = balance
    r["status"] = status
    return r


@api_router.get("/receivables")
async def list_receivables():
    docs = await db.receivables.find({}, {"_id": 0}).sort("due_date", 1).to_list(1000)
    return [receivable_view(d) for d in docs]


@api_router.post("/receivables")
async def create_receivable(payload: ReceivableCreate):
    obj = Receivable(**payload.model_dump())
    doc = obj.model_dump()
    await db.receivables.insert_one(doc)
    return receivable_view(clean(doc))


@api_router.put("/receivables/{rid}")
async def update_receivable(rid: str, body: dict):
    fields = {k: body[k] for k in ["client_name", "project", "total", "received", "due_date", "method"] if k in body}
    await db.receivables.update_one({"id": rid}, {"$set": fields})
    doc = await db.receivables.find_one({"id": rid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Conta não encontrada")
    return receivable_view(doc)


@api_router.post("/receivables/{rid}/payment")
async def register_payment(rid: str, body: dict):
    doc = await db.receivables.find_one({"id": rid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Conta não encontrada")
    amount = round(float(body.get("amount", 0) or 0), 2)
    if amount <= 0:
        raise HTTPException(400, "Valor inválido")
    method = body.get("method", doc.get("method", "")) or ""
    received = round((doc.get("received", 0) or 0) + amount, 2)
    total = doc.get("total", 0) or 0
    if received > total:
        received = total
    payment = {"amount": amount, "method": method, "date": now_iso()}
    await db.receivables.update_one({"id": rid}, {"$set": {"received": received, "method": method}, "$push": {"payments": payment}})
    doc = await db.receivables.find_one({"id": rid}, {"_id": 0})
    return receivable_view(doc)


@api_router.post("/receivables/{rid}/pay")
async def mark_receivable_paid(rid: str, body: dict = None):
    doc = await db.receivables.find_one({"id": rid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Conta não encontrada")
    total = doc.get("total", 0) or 0
    prev = doc.get("received", 0) or 0
    balance = round(total - prev, 2)
    method = (body or {}).get("method", doc.get("method", "")) or ""
    update = {"$set": {"received": total, "method": method}}
    if balance > 0:
        update["$push"] = {"payments": {"amount": balance, "method": method, "date": now_iso()}}
    await db.receivables.update_one({"id": rid}, update)
    doc = await db.receivables.find_one({"id": rid}, {"_id": 0})
    return receivable_view(doc)


@api_router.delete("/receivables/{rid}")
async def delete_receivable(rid: str):
    await db.receivables.delete_one({"id": rid})
    return {"ok": True}


# ---------------- Contas a Pagar (Payables) ----------------
def payable_view(p: dict):
    p["amount"] = round(max(p.get("amount", 0) or 0, 0), 2)
    st = p.get("status", "pendente")
    if st not in ("pago", "cancelado"):
        due = p.get("due_date", "") or ""
        overdue = False
        if due:
            try:
                overdue = date.fromisoformat(due[:10]) < today()
            except Exception:
                overdue = False
        st = "vencido" if overdue else "pendente"
    p["status"] = st
    return p


@api_router.get("/payables")
async def list_payables():
    docs = await db.payables.find({}, {"_id": 0}).sort("due_date", 1).to_list(1000)
    return [payable_view(d) for d in docs]


@api_router.get("/payables/categories")
async def payable_categories():
    return PAYABLE_CATEGORIES


@api_router.post("/payables")
async def create_payable(payload: PayableCreate):
    data = payload.model_dump()
    data["amount"] = max(data.get("amount", 0) or 0, 0)
    obj = Payable(**data)
    doc = obj.model_dump()
    await db.payables.insert_one(doc)
    return payable_view(clean(doc))


@api_router.put("/payables/{pid}")
async def update_payable(pid: str, body: dict):
    fields = {k: body[k] for k in ["supplier", "description", "category", "amount", "due_date", "paid_date", "method", "status", "notes"] if k in body}
    if "amount" in fields:
        fields["amount"] = max(float(fields["amount"] or 0), 0)
    if fields.get("status") == "pago" and not (fields.get("paid_date") or "").strip():
        existing = await db.payables.find_one({"id": pid}, {"_id": 0})
        if existing and not (existing.get("paid_date") or "").strip():
            fields["paid_date"] = now_iso()[:10]
    await db.payables.update_one({"id": pid}, {"$set": fields})
    doc = await db.payables.find_one({"id": pid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Conta não encontrada")
    return payable_view(doc)


@api_router.post("/payables/{pid}/pay")
async def mark_payable_paid(pid: str, body: dict = None):
    doc = await db.payables.find_one({"id": pid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Conta não encontrada")
    body = body or {}
    update = {"status": "pago", "paid_date": body.get("paid_date") or now_iso()[:10]}
    if body.get("method"):
        update["method"] = body["method"]
    await db.payables.update_one({"id": pid}, {"$set": update})
    doc = await db.payables.find_one({"id": pid}, {"_id": 0})
    return payable_view(doc)


@api_router.delete("/payables/{pid}")
async def delete_payable(pid: str):
    await db.payables.delete_one({"id": pid})
    return {"ok": True}


# ---------------- Loja: Categorias ----------------
@api_router.get("/store/categories")
async def list_store_categories():
    return await db.store_categories.find({}, {"_id": 0}).sort("name", 1).to_list(500)


@api_router.post("/store/categories")
async def create_store_category(body: dict):
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "O nome da categoria é obrigatório")
    existing = await db.store_categories.find_one({"name": name}, {"_id": 0})
    if existing:
        return existing
    obj = StoreCategory(name=name, description=body.get("description", ""), active=body.get("active", True))
    await db.store_categories.insert_one(obj.model_dump())
    return clean(obj.model_dump())


@api_router.put("/store/categories/{cid}")
async def update_store_category(cid: str, body: dict):
    fields = {k: body[k] for k in ["name", "description", "active"] if k in body}
    await db.store_categories.update_one({"id": cid}, {"$set": fields})
    doc = await db.store_categories.find_one({"id": cid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Categoria não encontrada")
    return doc


@api_router.patch("/store/categories/{cid}/toggle")
async def toggle_store_category(cid: str):
    doc = await db.store_categories.find_one({"id": cid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Categoria não encontrada")
    new_active = not doc.get("active", True)
    await db.store_categories.update_one({"id": cid}, {"$set": {"active": new_active}})
    doc["active"] = new_active
    return doc


@api_router.delete("/store/categories/{cid}")
async def delete_store_category(cid: str):
    await db.store_categories.delete_one({"id": cid})
    return {"ok": True}


# ---------------- Loja: Produtos ----------------
@api_router.get("/store/products")
async def list_products(category: str = "", active: str = ""):
    query = {}
    if category:
        query["category"] = category
    if active in ("true", "false"):
        query["active"] = active == "true"
    docs = await db.products.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return docs


@api_router.post("/store/products")
async def create_product(payload: ProductCreate):
    data = payload.model_dump()
    data["price"] = max(float(data.get("price", 0) or 0), 0)
    obj = Product(**data)
    doc = obj.model_dump()
    await db.products.insert_one(doc)
    return clean(doc)


@api_router.put("/store/products/{pid}")
async def update_product(pid: str, body: dict):
    fields = {k: body[k] for k in ["name", "description", "category", "price", "image_url", "sku", "active"] if k in body}
    if "price" in fields:
        fields["price"] = max(float(fields["price"] or 0), 0)
    await db.products.update_one({"id": pid}, {"$set": fields})
    doc = await db.products.find_one({"id": pid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Produto não encontrado")
    return doc


@api_router.patch("/store/products/{pid}/toggle")
async def toggle_product(pid: str):
    doc = await db.products.find_one({"id": pid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Produto não encontrado")
    new_active = not doc.get("active", True)
    await db.products.update_one({"id": pid}, {"$set": {"active": new_active}})
    doc["active"] = new_active
    return doc


@api_router.delete("/store/products/{pid}")
async def delete_product(pid: str):
    await db.products.delete_one({"id": pid})
    return {"ok": True}


# ---------------- Loja: Pedidos (Orders) ----------------
def order_total(items):
    return round(sum((i.get("price", 0) or 0) * (i.get("quantity", 1) or 1) for i in items), 2)


@api_router.get("/store/orders/states")
async def order_states():
    return ORDER_STATES


@api_router.get("/store/orders")
async def list_orders(status: str = ""):
    query = {}
    if status:
        query["status"] = status
    docs = await db.store_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return docs


@api_router.post("/store/orders")
async def create_order(payload: OrderCreate):
    data = payload.model_dump()
    items = data.get("items", [])
    count = await db.store_orders.count_documents({})
    number = f"ENC-{datetime.now().year}-{count + 1:04d}"
    status = data.get("status") or "novo"
    if status not in ORDER_STATES:
        status = "novo"
    obj = Order(number=number, customer_name=data.get("customer_name", ""), customer_email=data.get("customer_email", ""),
                customer_phone=data.get("customer_phone", ""),
                items=items, total=order_total(items), status=status, notes=data.get("notes", ""))
    doc = obj.model_dump()
    await db.store_orders.insert_one(doc)
    return clean(doc)


@api_router.put("/store/orders/{oid}")
async def update_order(oid: str, body: dict):
    fields = {k: body[k] for k in ["customer_name", "customer_email", "items", "notes", "status"] if k in body}
    if "items" in fields:
        fields["total"] = order_total(fields["items"])
    if "status" in fields and fields["status"] not in ORDER_STATES:
        raise HTTPException(400, "Estado inválido")
    await db.store_orders.update_one({"id": oid}, {"$set": fields})
    doc = await db.store_orders.find_one({"id": oid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Pedido não encontrado")
    return doc


@api_router.patch("/store/orders/{oid}/status")
async def update_order_status(oid: str, body: dict):
    status = body.get("status", "")
    if status not in ORDER_STATES:
        raise HTTPException(400, "Estado inválido")
    await db.store_orders.update_one({"id": oid}, {"$set": {"status": status}})
    doc = await db.store_orders.find_one({"id": oid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Pedido não encontrado")
    await order_history(oid, f"Estado operacional alterado para '{status}'.")
    if status == "cancelado":
        await create_notification("order_cancelled", "Pedido cancelado", f"Pedido {doc.get('number','')} cancelado", oid)
    elif status == "entregue":
        await create_notification("order_completed", "Pedido concluído", f"Pedido {doc.get('number','')} entregue", oid)
    return doc


@api_router.delete("/store/orders/{oid}")
async def delete_order(oid: str):
    await db.store_orders.delete_one({"id": oid})
    return {"ok": True}


# ---------------- Notificações & Atividades (Fase 4) ----------------
@api_router.get("/notifications")
async def list_notifications(limit: int = 50):
    return await db.notifications.find({}, {"_id": 0}).sort("created_at", -1).to_list(limit)


@api_router.get("/notifications/unread-count")
async def notifications_unread_count():
    return {"count": await db.notifications.count_documents({"read": False})}


@api_router.post("/notifications/{nid}/read")
async def notification_mark_read(nid: str):
    await db.notifications.update_one({"id": nid}, {"$set": {"read": True}})
    return {"ok": True}


@api_router.post("/notifications/read-all")
async def notifications_read_all():
    await db.notifications.update_many({"read": False}, {"$set": {"read": True}})
    return {"ok": True}


@api_router.get("/activities")
async def list_activities(limit: int = 30):
    return await db.activities.find({}, {"_id": 0}).sort("created_at", -1).to_list(limit)


# ---------------- Dashboard ----------------
@api_router.get("/dashboard/stats")
async def dashboard_stats():
    clients = await db.clients.find({}, {"_id": 0}).to_list(2000)
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(2000)
    galleries = await db.galleries.find({}, {"_id": 0}).to_list(2000)
    invoices = [invoice_totals(i) for i in await db.invoices.find({}, {"_id": 0}).to_list(2000)]

    now = datetime.now(timezone.utc)
    ym = now.strftime("%Y-%m")
    year = now.strftime("%Y")

    revenue_month = sum(i["total"] for i in invoices if i.get("status") == "paga" and i.get("issue_date", "").startswith(ym))
    revenue_year = sum(i["total"] for i in invoices if i.get("status") == "paga" and i.get("issue_date", "").startswith(year))
    pending = sum(i["total"] for i in invoices if i.get("status") == "pendente")

    active_clients = len([c for c in clients if c.get("status") == "ativo"])
    new_leads = len([c for c in clients if c.get("status") in ("lead", "potencial")])

    td = today()
    week_end = td + timedelta(days=7)
    sessions_week = len([s for s in sessions if s.get("date") and td.isoformat() <= s["date"] <= week_end.isoformat()])

    galleries_delivered = len([g for g in galleries if g.get("status") == "entregue"])
    galleries_pending = len([g for g in galleries if g.get("status") != "entregue"])

    # birthdays in next 30 days
    birthdays = []
    for c in clients:
        bd = c.get("birthdate")
        if not bd or len(bd) < 10:
            continue
        try:
            m, d = int(bd[5:7]), int(bd[8:10])
            this_year = td.replace(month=m, day=d)
            if this_year < td:
                this_year = this_year.replace(year=td.year + 1)
            days = (this_year - td).days
            if 0 <= days <= 30:
                birthdays.append({"name": c["name"], "photo": c.get("photo", ""), "date": this_year.isoformat(), "days": days})
        except Exception:
            pass
    birthdays.sort(key=lambda x: x["days"])

    # revenue per month (paid)
    months = {}
    for i in invoices:
        if i.get("status") == "paga":
            m = i.get("issue_date", "")[:7]
            if m:
                months[m] = months.get(m, 0) + i["total"]
    revenue_chart = [{"month": k, "value": round(v, 2)} for k, v in sorted(months.items())][-6:]

    # sessions per month
    smonths = {}
    for s in sessions:
        m = (s.get("date") or "")[:7]
        if m:
            smonths[m] = smonths.get(m, 0) + 1
    sessions_chart = [{"month": k, "value": v} for k, v in sorted(smonths.items())][-6:]

    # sales by service (session value grouped by type)
    services = {}
    for s in sessions:
        services[s.get("type", "outro")] = services.get(s.get("type", "outro"), 0) + (s.get("value") or 0)
    sales_by_service = [{"name": k, "value": round(v, 2)} for k, v in services.items()]

    # client origins
    origins = {}
    for c in clients:
        o = c.get("origin", "outro") or "outro"
        origins[o] = origins.get(o, 0) + 1
    client_origins = [{"name": k, "value": v} for k, v in origins.items()]

    upcoming = sorted([s for s in sessions if s.get("date", "") >= td.isoformat()], key=lambda x: x.get("date", ""))[:5]

    paid_orders_docs = await db.store_orders.find({"payment_status": "paid"}, {"_id": 0}).to_list(5000)
    paid_orders = len(paid_orders_docs)
    total_sales = round(sum(o.get("total", 0) or 0 for o in paid_orders_docs), 2)
    avg_ticket = round(total_sales / paid_orders, 2) if paid_orders else 0
    store_revenue_month = round(sum((o.get("total", 0) or 0) for o in paid_orders_docs if (o.get("paid_at") or "").startswith(ym)), 2)
    notifications_unread = await db.notifications.count_documents({"read": False})

    return {
        "revenue_month": round(revenue_month, 2),
        "revenue_year": round(revenue_year, 2),
        "active_clients": active_clients,
        "new_leads": new_leads,
        "sessions_week": sessions_week,
        "galleries_delivered": galleries_delivered,
        "galleries_pending": galleries_pending,
        "pending_payments": round(pending, 2),
        "birthdays": birthdays[:5],
        "revenue_chart": revenue_chart,
        "sessions_chart": sessions_chart,
        "sales_by_service": sales_by_service,
        "client_origins": client_origins,
        "upcoming_sessions": upcoming,
        "total_clients": len(clients),
        "total_sessions": len(sessions),
        "total_sales": total_sales,
        "paid_orders": paid_orders,
        "avg_ticket": avg_ticket,
        "store_revenue_month": store_revenue_month,
        "notifications_unread": notifications_unread,
    }


# ---------------- Seed ----------------
@api_router.post("/seed")
async def seed():
    _seed_org = await ensure_default_org()
    current_org.set(_seed_org["id"])
    if await db.clients.count_documents({}) > 0:
        return {"seeded": False, "message": "Dados já existem"}

    imgs = [
        "https://images.pexels.com/photos/7778884/pexels-photo-7778884.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940",
        "https://images.pexels.com/photos/7778888/pexels-photo-7778888.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940",
        "https://images.pexels.com/photos/23876288/pexels-photo-23876288.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940",
        "https://images.pexels.com/photos/5804239/pexels-photo-5804239.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940",
        "https://images.pexels.com/photos/13699196/pexels-photo-13699196.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940",
        "https://images.pexels.com/photos/8015871/pexels-photo-8015871.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940",
    ]
    avatar = "https://images.pexels.com/photos/36697538/pexels-photo-36697538.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=300&w=300"

    yr = datetime.now().year
    clients_data = [
        {"name": "Ana & Rui Ferreira", "email": "ana.rui@email.pt", "phone": "+351 912 345 678", "whatsapp": "+351 912 345 678", "address": "Rua das Flores 12, Sintra", "nif": "215678943", "birthdate": f"{yr-30}-07-22", "photo": avatar, "client_type": "casamento", "status": "ativo", "origin": "instagram", "tags": ["casamento", "premium"]},
        {"name": "Studio Belle Mode", "email": "geral@bellemode.pt", "phone": "+351 210 987 654", "whatsapp": "", "address": "Av. da Liberdade 200, Lisboa", "nif": "509887221", "birthdate": "", "client_type": "empresa", "status": "ativo", "origin": "recomendacao", "company": "Belle Mode", "tags": ["moda", "recorrente"], "favorite": True},
        {"name": "Marca Vinha do Sol", "email": "info@vinhadosol.pt", "phone": "+351 936 112 233", "address": "Quinta do Sol, Douro", "nif": "501223344", "client_type": "empresa", "status": "ativo", "origin": "website", "company": "Vinha do Sol", "tags": ["produto", "comercial"]},
        {"name": "Beatriz Costa", "email": "beatriz.c@email.pt", "phone": "+351 927 445 001", "birthdate": f"{yr-28}-06-30", "photo": avatar, "client_type": "particular", "status": "lead", "origin": "google", "tags": ["retrato"]},
        {"name": "João Marques", "email": "joao.m@email.pt", "phone": "+351 913 220 887", "birthdate": f"{yr-35}-08-05", "client_type": "particular", "status": "lead", "origin": "instagram", "tags": ["batizado"]},
    ]
    clients = [Client(**c) for c in clients_data]
    await db.clients.insert_many([c.model_dump() for c in clients])
    cmap = {c.name: c.id for c in clients}
    # Portal test credentials
    await db.clients.update_one({"email": "ana.rui@email.pt"}, {"$set": {"password_hash": hash_password("cliente123")}})
    await db.clients.update_one({"email": "beatriz.c@email.pt"}, {"$set": {"password_hash": hash_password("cliente123")}})

    def dt(offset):
        return (today() + timedelta(days=offset)).isoformat()

    sessions_data = [
        {"title": "Casamento Quinta dos Sonhos", "type": "casamento", "client_name": "Ana & Rui Ferreira", "date": dt(4), "time": "10:00", "location": "Sintra", "status": "confirmada", "photographer": "Estúdio", "value": 3200},
        {"title": "Editorial Primavera 2026", "type": "moda", "client_name": "Studio Belle Mode", "date": dt(2), "time": "14:00", "location": "Lisboa", "status": "agendada", "photographer": "Estúdio", "value": 1800},
        {"title": "Campanha Vinho Reserva", "type": "produto", "client_name": "Marca Vinha do Sol", "date": dt(-30), "time": "09:00", "location": "Estúdio", "status": "entregue", "photographer": "Estúdio", "value": 2400},
        {"title": "Retrato Corporativo", "type": "retrato", "client_name": "Beatriz Costa", "date": dt(9), "time": "16:00", "location": "Porto", "status": "agendada", "photographer": "Estúdio", "value": 450},
        {"title": "Batizado do Tomás", "type": "batizado", "client_name": "João Marques", "date": dt(-10), "time": "11:00", "location": "Braga", "status": "realizada", "photographer": "Estúdio", "value": 600},
    ]
    sessions = [Session(client_id=cmap.get(s["client_name"], ""), **s) for s in sessions_data]
    await db.sessions.insert_many([s.model_dump() for s in sessions])

    galleries_data = [
        {"title": "Casamento Ana & Rui — Seleção", "client_name": "Ana & Rui Ferreira", "cover": imgs[2], "status": "pendente",
         "photos": [Photo(url=imgs[2], name="cerimonia.jpg").model_dump(), Photo(url=imgs[3], name="detalhes.jpg").model_dump(), Photo(url=imgs[0], name="casal.jpg").model_dump()]},
        {"title": "Editorial Belle Mode", "client_name": "Studio Belle Mode", "cover": imgs[1], "status": "entregue",
         "photos": [Photo(url=imgs[1], name="look1.jpg").model_dump(), Photo(url=imgs[4], name="look2.jpg").model_dump()]},
        {"title": "Campanha Vinho Reserva", "client_name": "Marca Vinha do Sol", "cover": imgs[5], "status": "entregue",
         "photos": [Photo(url=imgs[5], name="garrafa.jpg").model_dump()]},
    ]
    galleries = [Gallery(**g) for g in galleries_data]
    await db.galleries.insert_many([g.model_dump() for g in galleries])

    events_data = [
        {"title": "Sessão Casamento Quinta", "client_name": "Ana & Rui Ferreira", "date": dt(4), "time": "10:00", "type": "casamento", "location": "Sintra"},
        {"title": "Editorial Primavera", "client_name": "Studio Belle Mode", "date": dt(2), "time": "14:00", "type": "moda", "location": "Lisboa"},
        {"title": "Reunião Beatriz Costa", "client_name": "Beatriz Costa", "date": dt(1), "time": "16:30", "type": "reuniao", "location": "Online"},
    ]
    events = [EventItem(**e) for e in events_data]
    await db.events.insert_many([e.model_dump() for e in events])

    inv1 = Invoice(number=f"{yr}-0001", client_name="Marca Vinha do Sol", type="fatura", service="produto", status="paga", due_date=dt(-15),
                   items=[InvoiceItem(description="Campanha produto — pacote completo", quantity=1, price=2400)], tax_rate=23)
    inv1.issue_date = dt(-30)
    inv2 = Invoice(number=f"{yr}-0002", client_name="Ana & Rui Ferreira", type="fatura", service="casamento", status="pendente", due_date=dt(20),
                   items=[InvoiceItem(description="Cobertura casamento", quantity=1, price=3200)], tax_rate=23)
    inv2.issue_date = dt(-2)
    inv3 = Invoice(number=f"{yr}-0003", client_name="João Marques", type="fatura", service="batizado", status="paga", due_date=dt(-5),
                   items=[InvoiceItem(description="Reportagem batizado", quantity=1, price=600)], tax_rate=23)
    inv3.issue_date = dt(-9)
    await db.invoices.insert_many([inv1.model_dump(), inv2.model_dump(), inv3.model_dump()])

    q1 = Quote(number=f"ORC-{yr}-0001", client_name="Beatriz Costa", title="Sessão de Retrato Corporativo",
               status="enviado", template="retrato", valid_until=dt(20),
               items=[InvoiceItem(description="Sessão de retrato (1h)", quantity=1, price=250),
                      InvoiceItem(description="10 fotografias editadas", quantity=1, price=100)], tax_rate=23)
    q2 = Quote(number=f"ORC-{yr}-0002", client_name="João Marques", title="Reportagem de Batizado",
               status="aprovado", template="personalizado", valid_until=dt(10),
               items=[InvoiceItem(description="Reportagem batizado", quantity=1, price=600)], tax_rate=23)
    await db.quotes.insert_many([q1.model_dump(), q2.model_dump()])

    ctr_body = CONTRACT_TEMPLATES[1]["body"].format(cliente="Ana & Rui Ferreira", titulo="Casamento Quinta dos Sonhos", valor="3 936,00 €", data=dt(-2))
    c1 = Contract(number=f"CTR-{yr}-0001", client_name="Ana & Rui Ferreira", title="Casamento Quinta dos Sonhos",
                  body=ctr_body, status="assinado", template="casamento", signer_name="Ana Ferreira", signed_at=now_iso())
    c2 = Contract(number=f"CTR-{yr}-0002", client_name="Studio Belle Mode", title="Editorial Primavera 2026",
                  body=CONTRACT_TEMPLATES[0]["body"].format(cliente="Studio Belle Mode", titulo="Editorial Primavera 2026", valor="2 214,00 €", data=dt(-1)),
                  status="enviado", template="servicos")
    await db.contracts.insert_many([c1.model_dump(), c2.model_dump()])

    payables = [
        {"id": str(uuid.uuid4()), "supplier": "Imobiliária Central", "description": "Renda do estúdio", "category": "Fornecedores", "amount": 650, "due_date": dt(-20), "paid_date": dt(-20), "method": "Transferência", "status": "pago", "notes": "", "created_at": now_iso()},
        {"id": str(uuid.uuid4()), "supplier": "Fnac Pro", "description": "Objetiva 85mm f/1.4", "category": "Equipamentos", "amount": 1200, "due_date": dt(-40), "paid_date": dt(-40), "method": "Cartão", "status": "pago", "notes": "", "created_at": now_iso()},
        {"id": str(uuid.uuid4()), "supplier": "Adobe", "description": "Creative Cloud (anual)", "category": "Software", "amount": 60, "due_date": dt(-5), "paid_date": dt(-5), "method": "Cartão", "status": "pago", "notes": "", "created_at": now_iso()},
        {"id": str(uuid.uuid4()), "supplier": "Galp", "description": "Combustível (deslocações)", "category": "Combustível", "amount": 180, "due_date": dt(-2), "paid_date": "", "method": "", "status": "pendente", "notes": "Sessões fora de Lisboa", "created_at": now_iso()},
        {"id": str(uuid.uuid4()), "supplier": "Meta Ads", "description": "Campanha Instagram", "category": "Publicidade", "amount": 120, "due_date": dt(-8), "paid_date": "", "method": "", "status": "pendente", "notes": "", "created_at": now_iso()},
        {"id": str(uuid.uuid4()), "supplier": "Gráfica Lumina", "description": "Impressões e álbum casamento", "category": "Álbuns", "amount": 300, "due_date": dt(-15), "paid_date": dt(-15), "method": "MB Way", "status": "pago", "notes": "", "created_at": now_iso()},
        {"id": str(uuid.uuid4()), "supplier": "Rita Design", "description": "Design de convites", "category": "Designers", "amount": 250, "due_date": dt(6), "paid_date": "", "method": "", "status": "pendente", "notes": "", "created_at": now_iso()},
        {"id": str(uuid.uuid4()), "supplier": "Autoridade Tributária", "description": "IVA trimestral", "category": "Impostos", "amount": 890, "due_date": dt(14), "paid_date": "", "method": "", "status": "pendente", "notes": "", "created_at": now_iso()},
    ]
    await db.payables.insert_many(payables)

    receivables = [
        {"id": str(uuid.uuid4()), "client_name": "Ana & Rui Ferreira", "project": "Casamento Quinta dos Sonhos", "total": 3936, "received": 1500, "due_date": dt(20), "method": "Transferência", "payments": [{"amount": 1500, "method": "Transferência", "date": now_iso()}], "created_at": now_iso()},
        {"id": str(uuid.uuid4()), "client_name": "Studio Belle Mode", "project": "Editorial Primavera 2026", "total": 2214, "received": 0, "due_date": dt(12), "method": "", "payments": [], "created_at": now_iso()},
        {"id": str(uuid.uuid4()), "client_name": "Marca Vinha do Sol", "project": "Campanha Vinho Reserva", "total": 2952, "received": 2952, "due_date": dt(-15), "method": "MB Way", "payments": [{"amount": 2952, "method": "MB Way", "date": now_iso()}], "created_at": now_iso()},
        {"id": str(uuid.uuid4()), "client_name": "Beatriz Costa", "project": "Retrato Corporativo", "total": 553.5, "received": 0, "due_date": dt(-6), "method": "", "payments": [], "created_at": now_iso()},
        {"id": str(uuid.uuid4()), "client_name": "João Marques", "project": "Batizado do Tomás", "total": 738, "received": 300, "due_date": dt(-3), "method": "Dinheiro", "payments": [{"amount": 300, "method": "Dinheiro", "date": now_iso()}], "created_at": now_iso()},
    ]
    await db.receivables.insert_many(receivables)

    store_categories = ["Impressões", "Álbuns", "Molduras", "Canvas", "Ficheiros Digitais", "Packs"]
    await db.store_categories.insert_many([StoreCategory(name=n).model_dump() for n in store_categories])

    products = [
        {"id": str(uuid.uuid4()), "name": "Impressão Fine Art 30x40", "description": "Impressão premium em papel Hahnemühle.", "category": "Impressões", "price": 35, "image_url": imgs[0], "sku": "IMP-3040", "active": True, "created_at": now_iso()},
        {"id": str(uuid.uuid4()), "name": "Álbum Luxo 30x30 (20 páginas)", "description": "Álbum capa dura com acabamento em linho.", "category": "Álbuns", "price": 180, "image_url": imgs[2], "sku": "ALB-3030", "active": True, "created_at": now_iso()},
        {"id": str(uuid.uuid4()), "name": "Moldura de Madeira 20x30", "description": "Moldura artesanal com vidro anti-reflexo.", "category": "Molduras", "price": 45, "image_url": imgs[3], "sku": "MOL-2030", "active": True, "created_at": now_iso()},
        {"id": str(uuid.uuid4()), "name": "Canvas 40x60", "description": "Tela esticada em bastidor de madeira.", "category": "Canvas", "price": 90, "image_url": imgs[4], "sku": "CAN-4060", "active": True, "created_at": now_iso()},
        {"id": str(uuid.uuid4()), "name": "Pack Digital — 10 fotos editadas", "description": "Ficheiros em alta resolução para download.", "category": "Ficheiros Digitais", "price": 60, "image_url": imgs[1], "sku": "DIG-PK10", "active": True, "created_at": now_iso()},
        {"id": str(uuid.uuid4()), "name": "Pack Casamento Completo", "description": "Álbum + 2 canvas + galeria digital.", "category": "Packs", "price": 450, "image_url": imgs[5], "sku": "PACK-WED", "active": False, "created_at": now_iso()},
    ]
    await db.products.insert_many(products)

    store_orders = [
        {"id": str(uuid.uuid4()), "number": "ENC-2026-0001", "customer_name": "Ana & Rui Ferreira", "customer_email": "ana.rui@email.pt",
         "items": [{"product_id": "", "name": "Álbum Luxo 30x30 (20 páginas)", "price": 180, "quantity": 1}, {"product_id": "", "name": "Canvas 40x60", "price": 90, "quantity": 2}],
         "total": 360, "status": "pago", "notes": "Entregar até ao fim do mês.", "created_at": now_iso()},
        {"id": str(uuid.uuid4()), "number": "ENC-2026-0002", "customer_name": "Beatriz Costa", "customer_email": "beatriz.c@email.pt",
         "items": [{"product_id": "", "name": "Impressão Fine Art 30x40", "price": 35, "quantity": 3}],
         "total": 105, "status": "novo", "notes": "", "created_at": now_iso()},
        {"id": str(uuid.uuid4()), "number": "ENC-2026-0003", "customer_name": "João Marques", "customer_email": "joao.m@email.pt",
         "items": [{"product_id": "", "name": "Moldura de Madeira 20x30", "price": 45, "quantity": 1}],
         "total": 45, "status": "enviado", "notes": "", "created_at": now_iso()},
    ]
    await db.store_orders.insert_many(store_orders)

    return {"seeded": True}


# ---------------- Templates ----------------
QUOTE_TEMPLATES = [
    {"id": "casamento", "name": "Casamento", "tax_rate": 23, "items": [
        {"description": "Cobertura fotográfica (8h)", "quantity": 1, "price": 1500},
        {"description": "Álbum premium 30x30", "quantity": 1, "price": 450},
        {"description": "Galeria online + downloads", "quantity": 1, "price": 150}]},
    {"id": "retrato", "name": "Sessão de Retrato", "tax_rate": 23, "items": [
        {"description": "Sessão de retrato (1h)", "quantity": 1, "price": 250},
        {"description": "10 fotografias editadas", "quantity": 1, "price": 100}]},
    {"id": "produto", "name": "Fotografia de Produto", "tax_rate": 23, "items": [
        {"description": "Sessão de produto (meio dia)", "quantity": 1, "price": 600},
        {"description": "Edição avançada por foto", "quantity": 10, "price": 15}]},
]
CONTRACT_TEMPLATES = [
    {"id": "servicos", "name": "Prestação de Serviços", "body": (
        "CONTRATO DE PRESTAÇÃO DE SERVIÇOS FOTOGRÁFICOS\n\n"
        "Entre o Estúdio StudioHub AI (Prestador) e {cliente} (Cliente).\n\n"
        "1. OBJETO — O Prestador compromete-se a realizar o serviço fotográfico \"{titulo}\".\n"
        "2. VALOR — O valor total acordado é de {valor}, com IVA incluído.\n"
        "3. DATA — O serviço será realizado na data acordada entre as partes.\n"
        "4. ENTREGA — As fotografias serão entregues em galeria online até 30 dias após a sessão.\n"
        "5. DIREITOS — O Prestador mantém os direitos de autor; o Cliente recebe licença de uso pessoal.\n\n"
        "Data: {data}\n\nAssinatura do Cliente: ______________________")},
    {"id": "casamento", "name": "Casamento", "body": (
        "CONTRATO DE COBERTURA DE CASAMENTO\n\n"
        "Entre o Estúdio StudioHub AI e {cliente}.\n\n"
        "1. EVENTO — Cobertura fotográfica do casamento \"{titulo}\".\n"
        "2. VALOR — {valor} (IVA incluído). Sinal de 30% na assinatura.\n"
        "3. ENTREGA — Galeria online em 45 dias; álbum em 90 dias.\n"
        "4. CANCELAMENTO — O sinal não é reembolsável.\n\n"
        "Data: {data}\n\nAssinatura do Cliente: ______________________")},
]


@api_router.get("/templates")
async def get_templates():
    return {"quotes": QUOTE_TEMPLATES, "contracts": CONTRACT_TEMPLATES}


# ---------------- Quotes ----------------
@api_router.get("/quotes")
async def list_quotes():
    docs = await db.quotes.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [invoice_totals(d) for d in docs]


@api_router.get("/quotes/{quote_id}")
async def get_quote(quote_id: str):
    doc = await db.quotes.find_one({"id": quote_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Orçamento não encontrado")
    return invoice_totals(doc)


@api_router.post("/quotes")
async def create_quote(payload: QuoteCreate):
    count = await db.quotes.count_documents({})
    number = f"ORC-{datetime.now().year}-{count + 1:04d}"
    obj = Quote(number=number, **payload.model_dump())
    doc = obj.model_dump()
    await db.quotes.insert_one(doc)
    return invoice_totals(clean(doc))


@api_router.put("/quotes/{quote_id}")
async def update_quote(quote_id: str, payload: QuoteCreate):
    if not await db.quotes.find_one({"id": quote_id}):
        raise HTTPException(404, "Orçamento não encontrado")
    await db.quotes.update_one({"id": quote_id}, {"$set": payload.model_dump()})
    return invoice_totals(await db.quotes.find_one({"id": quote_id}, {"_id": 0}))


@api_router.patch("/quotes/{quote_id}/status")
async def update_quote_status(quote_id: str, body: dict):
    await db.quotes.update_one({"id": quote_id}, {"$set": {"status": body.get("status", "rascunho")}})
    doc = await db.quotes.find_one({"id": quote_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Orçamento não encontrado")
    return invoice_totals(doc)


@api_router.delete("/quotes/{quote_id}")
async def delete_quote(quote_id: str):
    await db.quotes.delete_one({"id": quote_id})
    return {"ok": True}


@api_router.post("/quotes/{quote_id}/convert-to-invoice")
async def quote_to_invoice(quote_id: str):
    q = await db.quotes.find_one({"id": quote_id}, {"_id": 0})
    if not q:
        raise HTTPException(404, "Orçamento não encontrado")
    count = await db.invoices.count_documents({})
    inv = Invoice(number=f"{datetime.now().year}-{count + 1:04d}", client_name=q["client_name"],
                  items=[InvoiceItem(**i) for i in q.get("items", [])], tax_rate=q.get("tax_rate", 23),
                  notes=f"Gerada a partir do orçamento {q['number']}")
    await db.invoices.insert_one(inv.model_dump())
    await db.quotes.update_one({"id": quote_id}, {"$set": {"status": "convertido", "invoice_id": inv.id}})
    return invoice_totals(inv.model_dump())


@api_router.post("/quotes/{quote_id}/convert-to-contract")
async def quote_to_contract(quote_id: str):
    q = await db.quotes.find_one({"id": quote_id}, {"_id": 0})
    if not q:
        raise HTTPException(404, "Orçamento não encontrado")
    tot = invoice_totals(dict(q))
    tmpl = next((t for t in CONTRACT_TEMPLATES if t["id"] == "servicos"), CONTRACT_TEMPLATES[0])
    body = tmpl["body"].format(cliente=q["client_name"], titulo=q.get("title", "Serviço"),
                               valor=f"{tot['total']:.2f} €", data=today().isoformat())
    count = await db.contracts.count_documents({})
    c = Contract(number=f"CTR-{datetime.now().year}-{count + 1:04d}", client_name=q["client_name"],
                 title=q.get("title", "Contrato"), body=body, template="servicos", quote_id=quote_id)
    await db.contracts.insert_one(c.model_dump())
    await db.quotes.update_one({"id": quote_id}, {"$set": {"contract_id": c.id}})
    return c.model_dump()


# ---------------- Contracts ----------------
@api_router.get("/contracts")
async def list_contracts():
    return await db.contracts.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)


@api_router.get("/contracts/{contract_id}")
async def get_contract(contract_id: str):
    doc = await db.contracts.find_one({"id": contract_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Contrato não encontrado")
    return doc


@api_router.post("/contracts")
async def create_contract(payload: ContractCreate):
    count = await db.contracts.count_documents({})
    number = f"CTR-{datetime.now().year}-{count + 1:04d}"
    body = payload.body
    if not body and payload.template:
        tmpl = next((t for t in CONTRACT_TEMPLATES if t["id"] == payload.template), None)
        if tmpl:
            body = tmpl["body"].format(cliente=payload.client_name, titulo=payload.title, valor="a acordar", data=today().isoformat())
    obj = Contract(number=number, body=body, **payload.model_dump(exclude={"body"}))
    doc = obj.model_dump()
    await db.contracts.insert_one(doc)
    return clean(doc)


@api_router.put("/contracts/{contract_id}")
async def update_contract(contract_id: str, payload: ContractCreate):
    if not await db.contracts.find_one({"id": contract_id}):
        raise HTTPException(404, "Contrato não encontrado")
    await db.contracts.update_one({"id": contract_id}, {"$set": payload.model_dump()})
    return await db.contracts.find_one({"id": contract_id}, {"_id": 0})


@api_router.patch("/contracts/{contract_id}/status")
async def update_contract_status(contract_id: str, body: dict):
    await db.contracts.update_one({"id": contract_id}, {"$set": {"status": body.get("status", "rascunho")}})
    doc = await db.contracts.find_one({"id": contract_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Contrato não encontrado")
    return doc


@api_router.post("/contracts/{contract_id}/sign")
async def sign_contract(contract_id: str, body: dict):
    signer = body.get("signer_name", "").strip()
    if not signer:
        raise HTTPException(400, "Nome de assinatura obrigatório")
    await db.contracts.update_one({"id": contract_id}, {"$set": {"status": "assinado", "signer_name": signer, "signed_at": now_iso()}})
    doc = await db.contracts.find_one({"id": contract_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Contrato não encontrado")
    return doc


@api_router.delete("/contracts/{contract_id}")
async def delete_contract(contract_id: str):
    await db.contracts.delete_one({"id": contract_id})
    return {"ok": True}


# ---------------- AI Assistant ----------------
@api_router.get("/ai/history/{session_id}")
async def ai_history(session_id: str):
    return await db.ai_messages.find({"session_id": session_id}, {"_id": 0}).sort("ts", 1).to_list(200)


@api_router.post("/ai/chat")
async def ai_chat(payload: AiChatIn):
    session_id = payload.session_id or str(uuid.uuid4())

    clients = await db.clients.find({}, {"_id": 0}).to_list(500)
    invoices = [invoice_totals(i) for i in await db.invoices.find({}, {"_id": 0}).to_list(500)]
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(500)
    unpaid = [f"{i['client_name']} — {i['number']} ({i['total']:.2f} €)" for i in invoices if i.get("status") == "pendente"]
    upcoming = [f"{s['title']} — {s['client_name']} ({s.get('date','')})" for s in sessions if s.get("date", "") >= today().isoformat()][:8]

    context = (
        f"Dados do estúdio (usa apenas se relevante):\n"
        f"- Total de clientes: {len(clients)}\n"
        f"- Clientes: {', '.join(c['name'] for c in clients[:15])}\n"
        f"- Faturas por pagar: {'; '.join(unpaid) if unpaid else 'nenhuma'}\n"
        f"- Próximas sessões: {'; '.join(upcoming) if upcoming else 'nenhuma'}\n"
    )
    system = (
        "És o assistente inteligente do StudioHub AI, uma plataforma de gestão para fotógrafos. "
        "Ajudas o fotógrafo a gerir o negócio: redigir orçamentos, contratos, emails para clientes, "
        "campanhas de marketing, resumos e sugestões. Respondes SEMPRE em português de Portugal, "
        "de forma clara, profissional e concisa. Quando te pedirem para criar um orçamento, contrato ou email, "
        "produz o conteúdo pronto a usar, bem estruturado.\n\n" + context
    )

    # Replay recent conversation history for multi-turn continuity
    prior = await db.ai_messages.find({"session_id": session_id}, {"_id": 0}).sort("ts", 1).to_list(20)
    if prior:
        convo = "\n".join(f"{'Fotógrafo' if m['role'] == 'user' else 'Assistente'}: {m['content']}" for m in prior[-8:])
        system += f"\n\nHistórico recente da conversa:\n{convo}"

    try:
        chat = LlmChat(api_key=EMERGENT_LLM_KEY, session_id=f"assist-{session_id}", system_message=system).with_model("openai", "gpt-5.4")
        reply = await chat.send_message(UserMessage(text=payload.message))
        reply = reply if isinstance(reply, str) else str(reply)
    except Exception as e:
        logging.warning(f"AI chat failed: {e}")
        raise HTTPException(500, "O assistente não está disponível de momento.")

    ts = now_iso()
    await db.ai_messages.insert_many([
        {"session_id": session_id, "role": "user", "content": payload.message, "ts": ts},
        {"session_id": session_id, "role": "assistant", "content": reply, "ts": now_iso()},
    ])
    return {"session_id": session_id, "reply": reply}


@api_router.get("/settings")
async def get_settings():
    doc = await db.settings.find_one({"_key": "app"}, {"_id": 0, "_key": 0})
    if not doc:
        doc = dict(DEFAULT_SETTINGS)
        await db.settings.insert_one({"_key": "app", **doc})
    return doc


@api_router.put("/settings")
async def update_settings(payload: Settings):
    await db.settings.update_one({"_key": "app"}, {"$set": payload.model_dump()}, upsert=True)
    return payload.model_dump()


@api_router.get("/finance/summary")
async def finance_summary():
    invoices = [invoice_totals(i) for i in await db.invoices.find({}, {"_id": 0}).to_list(2000)]
    payables = [payable_view(p) for p in await db.payables.find({}, {"_id": 0}).to_list(2000)]
    recv_docs = await db.receivables.find({}, {"_id": 0}).to_list(2000)
    now = datetime.now(timezone.utc)
    ym, year = now.strftime("%Y-%m"), now.strftime("%Y")

    recv_open = sum(max((r.get("total", 0) or 0) - (r.get("received", 0) or 0), 0) for r in recv_docs)
    revenue_month = sum(i["total"] for i in invoices if i.get("status") == "paga" and i.get("issue_date", "").startswith(ym))
    receivable = sum(i["total"] for i in invoices if i.get("status") == "pendente") + recv_open
    payable = sum(p.get("amount", 0) for p in payables if p.get("status") in ("pendente", "vencido"))
    paid_rev_year = sum(i["total"] for i in invoices if i.get("status") == "paga" and i.get("issue_date", "").startswith(year))
    paid_exp_year = sum(p.get("amount", 0) for p in payables if p.get("status") == "pago" and ((p.get("paid_date") or p.get("due_date") or "")).startswith(year))
    total_paid_rev = sum(i["total"] for i in invoices if i.get("status") == "paga")
    total_paid_exp = sum(p.get("amount", 0) for p in payables if p.get("status") == "pago")

    months = {}
    for i in invoices:
        if i.get("status") == "paga":
            m = i.get("issue_date", "")[:7]
            if m:
                months[m] = months.get(m, 0) + i["total"]
    revenue_chart = [{"month": k, "value": round(v, 2)} for k, v in sorted(months.items())][-6:]

    cats = {}
    for p in payables:
        if p.get("status") == "pago":
            c = p.get("category", "Outros")
            cats[c] = cats.get(c, 0) + p.get("amount", 0)
    expenses_by_category = [{"name": k, "value": round(v, 2)} for k, v in sorted(cats.items(), key=lambda x: -x[1])]

    return {
        "revenue_month": round(revenue_month, 2),
        "receivable": round(receivable, 2),
        "payable": round(payable, 2),
        "profit": round(paid_rev_year - paid_exp_year, 2),
        "cashflow": round(total_paid_rev - total_paid_exp, 2),
        "revenue_chart": revenue_chart,
        "expenses_by_category": expenses_by_category,
    }


# ---------------- Relatórios Financeiros ----------------
def _in_range(d, start, end):
    if not d:
        return False
    d = d[:10]
    if start and d < start:
        return False
    if end and d > end:
        return False
    return True


async def build_financial_report(start="", end="", status="", category=""):
    invoices = [invoice_totals(i) for i in await db.invoices.find({}, {"_id": 0}).to_list(3000)]
    payables = [payable_view(p) for p in await db.payables.find({}, {"_id": 0}).to_list(3000)]
    receivables = [receivable_view(r) for r in await db.receivables.find({}, {"_id": 0}).to_list(3000)]

    # Receita por mês (faturas pagas)
    months = {}
    for i in invoices:
        if i.get("status") == "paga" and _in_range(i.get("issue_date", ""), start, end):
            m = i.get("issue_date", "")[:7]
            if m:
                months[m] = months.get(m, 0) + i["total"]
    revenue_by_month = [{"month": k, "value": round(v, 2)} for k, v in sorted(months.items())]

    # Despesas por categoria (pagas)
    cats = {}
    for p in payables:
        if p.get("status") == "pago" and _in_range(p.get("paid_date") or p.get("due_date") or "", start, end):
            if category and p.get("category") != category:
                continue
            c = p.get("category", "Outros")
            cats[c] = cats.get(c, 0) + p.get("amount", 0)
    expenses_by_category = [{"name": k, "value": round(v, 2)} for k, v in sorted(cats.items(), key=lambda x: -x[1])]

    # Contas a receber por estado
    r_status = {}
    for r in receivables:
        if not _in_range(r.get("due_date", ""), start, end):
            continue
        if status and r.get("status") != status:
            continue
        s = r.get("status", "pendente")
        r_status.setdefault(s, {"count": 0, "value": 0})
        r_status[s]["count"] += 1
        r_status[s]["value"] += r.get("balance", 0)
    receivables_by_status = [{"name": k, "count": v["count"], "value": round(v["value"], 2)} for k, v in r_status.items()]

    # Contas a pagar por estado
    p_status = {}
    for p in payables:
        if not _in_range(p.get("due_date", ""), start, end):
            continue
        if status and p.get("status") != status:
            continue
        if category and p.get("category") != category:
            continue
        s = p.get("status", "pendente")
        p_status.setdefault(s, {"count": 0, "value": 0})
        p_status[s]["count"] += 1
        p_status[s]["value"] += p.get("amount", 0)
    payables_by_status = [{"name": k, "count": v["count"], "value": round(v["value"], 2)} for k, v in p_status.items()]

    # Fluxo de caixa (entradas vs saídas) por mês
    flow = {}
    for i in invoices:
        if i.get("status") == "paga" and _in_range(i.get("issue_date", ""), start, end):
            m = i.get("issue_date", "")[:7]
            flow.setdefault(m, {"inflow": 0, "outflow": 0})
            flow[m]["inflow"] += i["total"]
    for p in payables:
        if p.get("status") == "pago" and _in_range(p.get("paid_date") or p.get("due_date") or "", start, end):
            m = (p.get("paid_date") or p.get("due_date") or "")[:7]
            if m:
                flow.setdefault(m, {"inflow": 0, "outflow": 0})
                flow[m]["outflow"] += p.get("amount", 0)
    cashflow = [{"month": k, "inflow": round(v["inflow"], 2), "outflow": round(v["outflow"], 2), "net": round(v["inflow"] - v["outflow"], 2)} for k, v in sorted(flow.items())]
    total_inflow = round(sum(c["inflow"] for c in cashflow), 2)
    total_outflow = round(sum(c["outflow"] for c in cashflow), 2)

    # Top 10 clientes por faturação
    clients = {}
    for i in invoices:
        if _in_range(i.get("issue_date", ""), start, end):
            n = i.get("client_name", "—")
            clients[n] = clients.get(n, 0) + i["total"]
    top_clients = [{"name": k, "value": round(v, 2)} for k, v in sorted(clients.items(), key=lambda x: -x[1])][:10]

    return {
        "filters": {"start": start, "end": end, "status": status, "category": category},
        "revenue_by_month": revenue_by_month,
        "expenses_by_category": expenses_by_category,
        "receivables_by_status": receivables_by_status,
        "payables_by_status": payables_by_status,
        "cashflow": cashflow,
        "totals": {"inflow": total_inflow, "outflow": total_outflow, "net": round(total_inflow - total_outflow, 2)},
        "top_clients": top_clients,
    }


@api_router.get("/reports/financial")
async def reports_financial(start: str = "", end: str = "", status: str = "", category: str = ""):
    return await build_financial_report(start, end, status, category)


STATUS_LABELS = {
    "paga": "Paga", "pendente": "Pendente", "cancelada": "Cancelada", "parcial": "Parcial",
    "pago": "Pago", "vencido": "Vencido", "cancelado": "Cancelado",
}


@api_router.get("/reports/financial/export")
async def export_financial_report(format: str = "xlsx", start: str = "", end: str = "", status: str = "", category: str = ""):
    data = await build_financial_report(start, end, status, category)
    period = f"{start or 'início'} a {end or 'hoje'}"

    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font
        wb = Workbook()

        def add_sheet(title, headers, rows):
            ws = wb.create_sheet(title[:31])
            ws.append(headers)
            for c in ws[1]:
                c.font = Font(bold=True)
            for r in rows:
                ws.append(r)

        wb.remove(wb.active)
        add_sheet("Receita por mês", ["Mês", "Receita (€)"], [[r["month"], r["value"]] for r in data["revenue_by_month"]])
        add_sheet("Despesas por categoria", ["Categoria", "Valor (€)"], [[r["name"], r["value"]] for r in data["expenses_by_category"]])
        add_sheet("Contas a receber", ["Estado", "Nº", "Saldo (€)"], [[STATUS_LABELS.get(r["name"], r["name"]), r["count"], r["value"]] for r in data["receivables_by_status"]])
        add_sheet("Contas a pagar", ["Estado", "Nº", "Valor (€)"], [[STATUS_LABELS.get(r["name"], r["name"]), r["count"], r["value"]] for r in data["payables_by_status"]])
        add_sheet("Fluxo de caixa", ["Mês", "Entradas (€)", "Saídas (€)", "Líquido (€)"], [[c["month"], c["inflow"], c["outflow"], c["net"]] for c in data["cashflow"]])
        add_sheet("Top 10 clientes", ["Cliente", "Faturação (€)"], [[r["name"], r["value"]] for r in data["top_clients"]])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=relatorio-financeiro.xlsx"})

    # PDF
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elems = [Paragraph("Relatório Financeiro — StudioHub AI", styles["Title"]),
             Paragraph(f"Período: {period}", styles["Normal"]), Spacer(1, 0.4 * cm)]

    def add_table(title, headers, rows):
        elems.append(Paragraph(title, styles["Heading2"]))
        table_data = [headers] + (rows if rows else [["—"] * len(headers)])
        t = Table(table_data, hAlign="LEFT")
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
            ("PADDING", (0, 0), (-1, -1), 4),
        ]))
        elems.append(t)
        elems.append(Spacer(1, 0.4 * cm))

    add_table("Receita por mês", ["Mês", "Receita (€)"], [[r["month"], f'{r["value"]:.2f}'] for r in data["revenue_by_month"]])
    add_table("Despesas por categoria", ["Categoria", "Valor (€)"], [[r["name"], f'{r["value"]:.2f}'] for r in data["expenses_by_category"]])
    add_table("Contas a receber por estado", ["Estado", "Nº", "Saldo (€)"], [[STATUS_LABELS.get(r["name"], r["name"]), str(r["count"]), f'{r["value"]:.2f}'] for r in data["receivables_by_status"]])
    add_table("Contas a pagar por estado", ["Estado", "Nº", "Valor (€)"], [[STATUS_LABELS.get(r["name"], r["name"]), str(r["count"]), f'{r["value"]:.2f}'] for r in data["payables_by_status"]])
    add_table("Fluxo de caixa", ["Mês", "Entradas", "Saídas", "Líquido"], [[c["month"], f'{c["inflow"]:.2f}', f'{c["outflow"]:.2f}', f'{c["net"]:.2f}'] for c in data["cashflow"]])
    add_table("Top 10 clientes por faturação", ["Cliente", "Faturação (€)"], [[r["name"], f'{r["value"]:.2f}'] for r in data["top_clients"]])
    doc.build(elems)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=relatorio-financeiro.pdf"})



@api_router.get("/")
async def root():
    return {"message": "StudioHub AI API"}


# ================= CLIENT PORTAL (Área do Cliente) =================
class PortalLogin(BaseModel):
    email: str
    password: str


class PortalForgot(BaseModel):
    email: str


class PortalProfile(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    tax_id: Optional[str] = None
    address: Optional[str] = None
    postal_code: Optional[str] = None
    city: Optional[str] = None
    region: Optional[str] = None


def _public_client(doc: dict) -> dict:
    doc = dict(doc)
    doc.pop("_id", None)
    doc.pop("password_hash", None)
    return doc


async def get_current_client(authorization: Optional[str] = Header(None)) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Não autenticado")
    token = authorization[7:]
    try:
        payload = pyjwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(401, "Sessão expirada")
    except pyjwt.InvalidTokenError:
        raise HTTPException(401, "Token inválido")
    client = await db.clients.find_one({"id": payload.get("sub")})
    if not client:
        raise HTTPException(401, "Cliente não encontrado")
    return client


@api_router.post("/portal/auth/login")
async def portal_login(payload: PortalLogin):
    email = payload.email.strip().lower()
    client = await db.clients.find_one({"email": {"$regex": f"^{email}$", "$options": "i"}})
    if not client or not client.get("password_hash") or not verify_password(payload.password, client["password_hash"]):
        raise HTTPException(401, "Email ou palavra-passe incorretos")
    token = create_client_token(client["id"], client.get("email", ""))
    return {"token": token, "client": _public_client(client)}


@api_router.get("/portal/auth/me")
async def portal_me(client: dict = Depends(get_current_client)):
    return _public_client(client)


@api_router.post("/portal/auth/forgot-password")
async def portal_forgot(payload: PortalForgot):
    email = payload.email.strip().lower()
    client = await db.clients.find_one({"email": {"$regex": f"^{email}$", "$options": "i"}})
    if client:
        token = secrets.token_urlsafe(32)
        await db.password_reset_tokens.insert_one({
            "client_id": client["id"], "token": token, "used": False,
            "expires_at": datetime.now(timezone.utc) + timedelta(hours=1)})
        logging.info(f"[PORTAL RESET] Link para {email}: /portal/reset?token={token}")
    return {"ok": True, "message": "Se o email existir, enviámos instruções de recuperação."}


@api_router.get("/portal/dashboard")
async def portal_dashboard(client: dict = Depends(get_current_client)):
    name = client.get("name", "")
    sessions = await db.sessions.find({"client_name": name}, {"_id": 0}).sort("date", 1).to_list(200)
    galleries = await db.galleries.find({"client_name": name}, {"_id": 0}).sort("created_at", -1).to_list(200)
    contracts = await db.contracts.find({"client_name": name}, {"_id": 0}).sort("created_at", -1).to_list(50)
    invoices = [invoice_totals(i) for i in await db.invoices.find({"client_name": name}, {"_id": 0}).to_list(200)]
    td = today().isoformat()
    upcoming = [s for s in sessions if s.get("date", "") >= td]
    pending = sum(i["total"] for i in invoices if i.get("status") == "pendente")
    paid = sum(i["total"] for i in invoices if i.get("status") == "paga")
    return {
        "client": _public_client(client),
        "next_session": upcoming[0] if upcoming else None,
        "galleries": galleries[:4],
        "documents": contracts[:4],
        "pending_payments": round(pending, 2),
        "paid_total": round(paid, 2),
        "counts": {"sessions": len(sessions), "galleries": len(galleries), "invoices": len(invoices)},
    }


@api_router.get("/portal/sessions")
async def portal_sessions(client: dict = Depends(get_current_client)):
    return await db.sessions.find({"client_name": client.get("name", "")}, {"_id": 0}).sort("date", -1).to_list(500)


@api_router.get("/portal/galleries")
async def portal_galleries(client: dict = Depends(get_current_client)):
    return await db.galleries.find({"client_name": client.get("name", "")}, {"_id": 0}).sort("created_at", -1).to_list(500)


@api_router.get("/portal/contracts")
async def portal_contracts(client: dict = Depends(get_current_client)):
    return await db.contracts.find({"client_name": client.get("name", "")}, {"_id": 0}).sort("created_at", -1).to_list(500)


@api_router.get("/portal/quotes")
async def portal_quotes(client: dict = Depends(get_current_client)):
    docs = await db.quotes.find({"client_name": client.get("name", "")}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [invoice_totals(d) for d in docs]


@api_router.get("/portal/invoices")
async def portal_invoices(client: dict = Depends(get_current_client)):
    docs = await db.invoices.find({"client_name": client.get("name", "")}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [invoice_totals(d) for d in docs]


@api_router.put("/portal/profile")
async def portal_profile(payload: PortalProfile, client: dict = Depends(get_current_client)):
    upd = {k: v for k, v in payload.model_dump().items() if v is not None}
    if upd:
        await db.clients.update_one({"id": client["id"]}, {"$set": upd})
    doc = await db.clients.find_one({"id": client["id"]})
    return _public_client(doc)


# ---------------- Fase 5.0: Autenticação do Fotógrafo ----------------
DEFAULT_ORG_NAME = "Studio E Fotografias"


class RegisterInput(BaseModel):
    name: str = ""
    email: str
    password: str


class LoginInput(BaseModel):
    email: str
    password: str


def create_studio_token(user_id: str, email: str, organization_id: str = "") -> str:
    payload = {"sub": user_id, "email": email, "role": "studio", "organization_id": organization_id,
               "exp": datetime.now(timezone.utc) + timedelta(days=7)}
    return pyjwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


async def ensure_default_org() -> dict:
    org = await db.organizations.find_one({"name": DEFAULT_ORG_NAME}, {"_id": 0})
    if not org:
        org = {"id": str(uuid.uuid4()), "name": DEFAULT_ORG_NAME, "slug": "studio-e-fotografias",
               "status": "active", "created_at": now_iso()}
        await db.organizations.insert_one(dict(org))
    return org


async def ensure_membership(user_id: str, organization_id: str, role: str = "owner"):
    existing = await db.organization_members.find_one({"user_id": user_id, "organization_id": organization_id})
    if not existing:
        await db.organization_members.insert_one({
            "id": str(uuid.uuid4()), "organization_id": organization_id, "user_id": user_id,
            "role": role, "created_at": now_iso()})


def _public_user(u: dict) -> dict:
    return {"id": u.get("id", ""), "name": u.get("name", ""), "email": u.get("email", ""),
            "organization_id": u.get("organization_id", ""), "role": u.get("role", "owner")}


async def get_current_user(authorization: Optional[str] = Header(None)) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Não autenticado")
    token = authorization.split(" ", 1)[1]
    try:
        payload = pyjwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(401, "Sessão expirada")
    except pyjwt.InvalidTokenError:
        raise HTTPException(401, "Token inválido")
    if payload.get("role") != "studio":
        raise HTTPException(401, "Token inválido")
    u = await db.users.find_one({"id": payload.get("sub")}, {"_id": 0})
    if not u:
        raise HTTPException(401, "Utilizador não encontrado")
    return u


@api_router.post("/auth/register")
async def auth_register(payload: RegisterInput):
    email = payload.email.strip().lower()
    if not email or not payload.password:
        raise HTTPException(400, "Email e palavra-passe obrigatórios")
    if await db.users.find_one({"email": email}):
        raise HTTPException(409, "Já existe uma conta com este email")
    org = await ensure_default_org()
    user = {"id": str(uuid.uuid4()), "name": (payload.name or "").strip() or "Fotógrafo", "email": email,
            "password_hash": hash_password(payload.password), "organization_id": org["id"],
            "role": "owner", "created_at": now_iso()}
    await db.users.insert_one(dict(user))
    await ensure_membership(user["id"], org["id"], "owner")
    return {"token": create_studio_token(user["id"], email, org["id"]), "user": _public_user(user)}


@api_router.post("/auth/login")
async def auth_login(payload: LoginInput):
    email = payload.email.strip().lower()
    u = await db.users.find_one({"email": email})
    if not u or not verify_password(payload.password, u.get("password_hash", "")):
        raise HTTPException(401, "Credenciais inválidas")
    return {"token": create_studio_token(u["id"], email, u.get("organization_id", "")), "user": _public_user(u)}


@api_router.get("/auth/me")
async def auth_me(user: dict = Depends(get_current_user)):
    return _public_user(user)


@api_router.post("/auth/logout")
async def auth_logout(user: dict = Depends(get_current_user)):
    return {"ok": True}


@api_router.post("/auth/forgot-password")
async def auth_forgot(body: dict):
    email = (body.get("email") or "").strip().lower()
    u = await db.users.find_one({"email": email})
    if u:
        tok = secrets.token_urlsafe(32)
        await db.password_reset_tokens.insert_one({
            "token": tok, "user_id": u["id"], "scope": "studio", "used": False,
            "expires_at": (datetime.now(timezone.utc) + timedelta(hours=1)).isoformat(),
            "created_at": now_iso()})
        logging.info(f"[STUDIO RESET] Link para {email}: /reset-password?token={tok}")
    return {"ok": True}


@api_router.post("/auth/reset-password")
async def auth_reset(body: dict):
    tok = body.get("token", ""); new_pw = body.get("password", "")
    if not tok or not new_pw:
        raise HTTPException(400, "Dados inválidos")
    rec = await db.password_reset_tokens.find_one({"token": tok, "scope": "studio"})
    if not rec or rec.get("used"):
        raise HTTPException(400, "Token inválido ou já usado")
    try:
        if datetime.fromisoformat(rec["expires_at"]) < datetime.now(timezone.utc):
            raise HTTPException(400, "Token expirado")
    except HTTPException:
        raise
    except Exception:
        pass
    await db.users.update_one({"id": rec["user_id"]}, {"$set": {"password_hash": hash_password(new_pw)}})
    await db.password_reset_tokens.update_one({"token": tok}, {"$set": {"used": True}})
    return {"ok": True}


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@app.on_event("startup")
async def _startup_auth():
    org = await ensure_default_org()
    admin_email = os.environ.get("ADMIN_EMAIL", "").strip().lower()
    admin_pw = os.environ.get("ADMIN_PASSWORD", "")
    if admin_email and admin_pw:
        existing = await db.users.find_one({"email": admin_email})
        if not existing:
            await db.users.insert_one({
                "id": str(uuid.uuid4()), "name": DEFAULT_ORG_NAME, "email": admin_email,
                "password_hash": hash_password(admin_pw), "organization_id": org["id"],
                "role": "owner", "created_at": now_iso()})
    # Backfill idempotente: todos os utilizadores ficam ligados a uma organização (membership)
    async for u in db.users.find({}):
        oid = u.get("organization_id") or org["id"]
        if not u.get("organization_id"):
            await db.users.update_one({"id": u["id"]}, {"$set": {"organization_id": oid}})
        await ensure_membership(u["id"], oid, u.get("role", "owner"))
    try:
        await db.users.create_index("email", unique=True)
        await db.organization_members.create_index([("organization_id", 1), ("user_id", 1)], unique=True)
    except Exception:
        pass
    # Migração idempotente: dados de negócio existentes ficam na organização padrão
    for coll in SCOPED_COLLECTIONS:
        await _raw_db[coll].update_many(
            {"organization_id": {"$exists": False}}, {"$set": {"organization_id": org["id"]}})


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
    client.close()
