"""Iteration 10 — Módulo Relatórios Financeiros
Testa GET /api/reports/financial (com filtros) e GET /api/reports/financial/export (xlsx/pdf).
Também valida regressão em endpoints de /api/finance/summary, /api/invoices, /api/receivables, /api/payables.
"""
import os
import io
import pytest
import requests

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
API = f"{BASE_URL}/api"


# ---------- helpers ----------
def _get(path, **params):
    r = requests.get(f"{API}{path}", params=params, timeout=30)
    return r


# ---------- Report JSON shape ----------
class TestReportShape:
    def test_report_returns_all_top_level_keys(self):
        r = _get("/reports/financial")
        assert r.status_code == 200, r.text
        j = r.json()
        expected = {
            "filters", "revenue_by_month", "expenses_by_category",
            "receivables_by_status", "payables_by_status",
            "cashflow", "totals", "top_clients",
        }
        assert expected.issubset(j.keys()), f"missing keys: {expected - set(j.keys())}"

    def test_totals_has_inflow_outflow_net(self):
        j = _get("/reports/financial").json()
        assert set(j["totals"].keys()) == {"inflow", "outflow", "net"}
        # net = inflow - outflow (rounded)
        assert abs(round(j["totals"]["inflow"] - j["totals"]["outflow"] - j["totals"]["net"], 2)) < 0.01

    def test_top_clients_max_10(self):
        j = _get("/reports/financial").json()
        assert isinstance(j["top_clients"], list)
        assert len(j["top_clients"]) <= 10
        for c in j["top_clients"]:
            assert "name" in c and "value" in c
            assert isinstance(c["value"], (int, float))

    def test_cashflow_items_shape(self):
        j = _get("/reports/financial").json()
        for c in j["cashflow"]:
            assert set(c.keys()) == {"month", "inflow", "outflow", "net"}
            assert abs(round(c["inflow"] - c["outflow"] - c["net"], 2)) < 0.01

    def test_revenue_by_month_shape(self):
        j = _get("/reports/financial").json()
        for r in j["revenue_by_month"]:
            assert "month" in r and "value" in r
            assert isinstance(r["value"], (int, float))

    def test_expenses_by_category_shape(self):
        j = _get("/reports/financial").json()
        for r in j["expenses_by_category"]:
            assert "name" in r and "value" in r

    def test_receivables_and_payables_by_status_shape(self):
        j = _get("/reports/financial").json()
        for r in j["receivables_by_status"]:
            assert set(r.keys()) == {"name", "count", "value"}
        for r in j["payables_by_status"]:
            assert set(r.keys()) == {"name", "count", "value"}

    def test_filters_echoed(self):
        j = _get("/reports/financial", start="2026-07-01", end="2026-07-31",
                 status="pendente", category="Marketing").json()
        assert j["filters"] == {"start": "2026-07-01", "end": "2026-07-31",
                                "status": "pendente", "category": "Marketing"}


# ---------- Filter behaviour ----------
class TestReportFilters:
    def test_period_filter_reduces_or_keeps_revenue(self):
        """Aplicar um período estrito deve reduzir (ou manter) o nº de meses de revenue."""
        base = _get("/reports/financial").json()
        july = _get("/reports/financial", start="2026-07-01", end="2026-07-31").json()
        base_months = {r["month"] for r in base["revenue_by_month"]}
        july_months = {r["month"] for r in july["revenue_by_month"]}
        # Every month in july filter must start with 2026-07
        for m in july_months:
            assert m.startswith("2026-07"), f"unexpected month {m}"
        # filtered set is subset of base
        assert july_months.issubset(base_months)

    def test_period_filter_excludes_outside_data(self):
        """Um período fora dos dados deve devolver estruturas vazias."""
        j = _get("/reports/financial", start="1990-01-01", end="1990-12-31").json()
        assert j["revenue_by_month"] == []
        assert j["cashflow"] == []
        assert j["totals"]["inflow"] == 0
        assert j["totals"]["outflow"] == 0
        assert j["totals"]["net"] == 0
        assert j["top_clients"] == []

    def test_status_filter_restricts_receivables_and_payables(self):
        j = _get("/reports/financial", status="pago").json()
        for r in j["receivables_by_status"]:
            assert r["name"] == "pago"
        for p in j["payables_by_status"]:
            assert p["name"] == "pago"

    def test_status_filter_pendente(self):
        j = _get("/reports/financial", status="pendente").json()
        for r in j["receivables_by_status"]:
            assert r["name"] == "pendente"
        for p in j["payables_by_status"]:
            assert p["name"] == "pendente"

    def test_category_filter_restricts_expenses_and_payables(self):
        j = _get("/reports/financial", category="Marketing").json()
        # expenses_by_category (built from pagos) — só Marketing
        for e in j["expenses_by_category"]:
            assert e["name"] == "Marketing"
        # payables_by_status vem agrupado por status; se filtrado por categoria,
        # os valores devem apenas refletir contas de Marketing. Não podemos ler categoria
        # no bucket agregado, mas ao filtrar por uma categoria inexistente deve ficar vazio.

    def test_category_filter_nonexistent_yields_empty_expenses(self):
        j = _get("/reports/financial", category="__CAT_INEXISTENTE__").json()
        assert j["expenses_by_category"] == []
        assert j["payables_by_status"] == []


# ---------- Export endpoints ----------
class TestReportExport:
    def test_export_xlsx_returns_spreadsheet(self):
        r = requests.get(f"{API}/reports/financial/export", params={"format": "xlsx"}, timeout=30)
        assert r.status_code == 200
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct, ct
        cd = r.headers.get("content-disposition", "")
        assert "attachment" in cd.lower()
        assert cd.lower().endswith(".xlsx") or ".xlsx" in cd.lower()
        # xlsx is a zip — starts with PK
        assert r.content[:2] == b"PK", "not a valid xlsx (missing PK header)"
        assert len(r.content) > 200

    def test_export_pdf_returns_pdf(self):
        r = requests.get(f"{API}/reports/financial/export", params={"format": "pdf"}, timeout=30)
        assert r.status_code == 200
        ct = r.headers.get("content-type", "")
        assert "application/pdf" in ct, ct
        cd = r.headers.get("content-disposition", "")
        assert "attachment" in cd.lower()
        assert ".pdf" in cd.lower()
        assert r.content[:4] == b"%PDF", "invalid pdf header"
        assert len(r.content) > 200

    def test_export_accepts_all_filters(self):
        r = requests.get(f"{API}/reports/financial/export",
                         params={"format": "xlsx", "start": "2026-07-01", "end": "2026-07-31",
                                 "status": "pendente", "category": "Marketing"},
                         timeout=30)
        assert r.status_code == 200
        assert r.content[:2] == b"PK"

    def test_export_pdf_accepts_filters(self):
        r = requests.get(f"{API}/reports/financial/export",
                         params={"format": "pdf", "start": "2026-07-01", "end": "2026-07-31"},
                         timeout=30)
        assert r.status_code == 200
        assert r.content[:4] == b"%PDF"

    def test_export_xlsx_is_openable(self):
        """valida que o xlsx abre no openpyxl e contém sheets esperados."""
        r = requests.get(f"{API}/reports/financial/export", params={"format": "xlsx"}, timeout=30)
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        expected = {"Receita por mês", "Despesas por categoria", "Contas a receber",
                    "Contas a pagar", "Fluxo de caixa", "Top 10 clientes"}
        got = set(wb.sheetnames)
        assert expected.issubset(got), f"missing sheets: {expected - got}"


# ---------- Regressão de módulos anteriores ----------
class TestRegression:
    def test_finance_summary_ok(self):
        r = _get("/finance/summary")
        assert r.status_code == 200
        j = r.json()
        for k in ["revenue_month", "receivable", "payable", "profit", "cashflow"]:
            assert k in j, f"kpi {k} missing"

    def test_invoices_list_ok(self):
        r = _get("/invoices")
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_receivables_list_ok(self):
        r = _get("/receivables")
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_payables_list_ok(self):
        r = _get("/payables")
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_clients_list_ok(self):
        r = _get("/clients")
        assert r.status_code == 200

    def test_galleries_list_ok(self):
        r = _get("/galleries")
        assert r.status_code == 200

    def test_events_list_ok(self):
        r = _get("/events")
        assert r.status_code == 200
